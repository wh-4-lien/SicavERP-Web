# vistas/inventario.py
import flet as ft
from core.database import get_sb, fetch_all
from core.estado import estado
from core.utilidades import hilo, clp


def _label_bodega(b: dict) -> str:
    area = (b.get("areas") or {}).get("nombre") or ""
    return f"[{area}] {b['nombre']}" if area else b["nombre"]

def badge_stock(stock: int) -> ft.Container:
    color = "red600" if stock == 0 else "orange600" if stock <= 3 else "green700"
    return ft.Container(
        content=ft.Text(str(stock), size=11, weight="bold", color="white"),
        bgcolor=color, border_radius=12,
        padding=ft.Padding(10, 2, 10, 2),
    )

class VistaInventario(ft.Container):
    def __init__(self, page: ft.Page):
        super().__init__()
        self.page_ref = page
        self.expand = True
        self.datos_cargados = False
        self._cargando = False

        self._productos_base = []
        self._prods_filtrados = []
        self._stock_bodega = {}
        self._sku_bodegas_map = {}
        self._fam_actual = "Todas"
        self._subfam_actual = "Todas"
        self._arbol_data = {}
        self._arbol_expandido = set()
        self._pagina = 0
        self._por_pagina = 25

        self.construir_ui()
        self.content = self._main_content

    # ══════════════════════════════ UI ══════════════════════════════
    def construir_ui(self):
        self.txt_filtro = ft.Text("Mostrando: Todas", size=11, italic=True, color="blue700")

        self.in_buscar = ft.TextField(
            label="Buscar por SKU o nombre", width=240, height=55,
            prefix_icon=ft.Icons.SEARCH,
            on_submit=lambda e: self._cargar_tabla(),
        )
        self.in_buscar.on_change = self._buscar_preview

        self._txt_preview = ft.Text("", size=12, weight="bold")
        self._cont_preview = ft.Container(
            content=self._txt_preview,
            visible=False,
            padding=ft.Padding(12, 6, 12, 6),
            border_radius=8,
            bgcolor="blue50",
            border=ft.Border(left=ft.BorderSide(3, "blue400")),
        )

        # Dropdown de bodega: on_change asignado después (Flet 0.85.x)
        self.dd_bodega = ft.Dropdown(label="Bodega", width=210, height=55,
                                     options=[ft.dropdown.Option("0", "Global (Todas)")])
        self.dd_bodega.value = "0"
        self.dd_bodega.on_change = self._cambiar_bodega

        self._ver_categorias = estado.puede_ver("categorias")

        cols = [ft.DataColumn(ft.Text("SKU", weight="bold", size=12))]
        if self._ver_categorias:
            cols += [
                ft.DataColumn(ft.Text("Familia",    weight="bold", size=12)),
                ft.DataColumn(ft.Text("Subfamilia", weight="bold", size=12, color="blue700")),
            ]
        cols += [
            ft.DataColumn(ft.Text("Producto",  weight="bold", size=12)),
            ft.DataColumn(ft.Text("Bodega / Área", weight="bold", size=12)),
            ft.DataColumn(ft.Text("Stock",     weight="bold", size=12), numeric=True),
            ft.DataColumn(ft.Text("Venta ($)", weight="bold", size=12), numeric=True),
        ]
        if estado.puede_ver("costo"):
            cols.append(ft.DataColumn(ft.Text("Costo ($)", weight="bold", size=12, color="red700"), numeric=True))

        self.tabla = ft.DataTable(
            columns=cols, rows=[],
            column_spacing=20, heading_row_height=40, data_row_max_height=38,
            divider_thickness=0.5,
            heading_row_color=ft.Colors.with_opacity(0.04, "blue"),
        )

        self.lbl_pagina = ft.Text("", size=12, color="grey700")
        self.btn_prev = ft.IconButton(
            ft.Icons.CHEVRON_LEFT, tooltip="Página anterior",
            on_click=lambda _: self._ir_pagina(self._pagina - 1),
        )
        self.btn_next = ft.IconButton(
            ft.Icons.CHEVRON_RIGHT, tooltip="Página siguiente",
            on_click=lambda _: self._ir_pagina(self._pagina + 1),
        )
        self.in_ir_pagina = ft.TextField(
            label="Ir a pág.", width=90, height=40,
            text_align="center", keyboard_type=ft.KeyboardType.NUMBER,
        )
        self.in_ir_pagina.on_submit = self._saltar_pagina
        self.row_paginacion = ft.Row(
            [
                self.btn_prev, self.lbl_pagina, self.btn_next,
                ft.Container(width=16),
                self.in_ir_pagina,
                ft.IconButton(ft.Icons.ARROW_FORWARD, tooltip="Ir", on_click=self._saltar_pagina),
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        )

        self.col_arbol = ft.Column([], spacing=2, scroll=ft.ScrollMode.AUTO)

        from utils.ui_helpers import sandwich

        _header = ft.Row([
            ft.Column([
                ft.Text("Inventario General", size=22, weight="bold", color="grey900"),
                self.txt_filtro,
            ], spacing=2, expand=True),
            ft.Container(content=ft.Row([
                self.dd_bodega,
                self.in_buscar,
                ft.Container(
                    content=ft.Row([ft.Icon(ft.Icons.SEARCH, color="white", size=16),
                                    ft.Text("Buscar", color="white", size=13, weight="bold")],
                                   spacing=6, tight=True),
                    bgcolor="blue700", border_radius=8,
                    padding=ft.Padding(14, 10, 14, 10), ink=True,
                    on_click=lambda _: self._cargar_tabla(),
                    tooltip="Buscar productos",
                ),
                ft.Container(
                    content=ft.Text("✕ Limpiar", color="grey600", size=13),
                    padding=ft.Padding(10, 10, 10, 10), border_radius=8, ink=True,
                    on_click=self._limpiar,
                    tooltip="Limpiar filtros",
                ),
                ft.IconButton(ft.Icons.REFRESH, tooltip="Actualizar inventario", on_click=self._recargar),
                ft.Container(width=8),
                ft.PopupMenuButton(
                    icon=ft.Icons.DOWNLOAD,
                    tooltip="Exportar inventario",
                    visible=estado.puede_hacer("inventario", "exportar"),
                    items=[
                        ft.PopupMenuItem(
                            content=ft.Row([ft.Icon(ft.Icons.TABLE_CHART, size=18), ft.Text("Exportar inventario global")], spacing=8),
                            on_click=lambda _: self._exportar("global"),
                        ),
                        ft.PopupMenuItem(
                            content=ft.Row([ft.Icon(ft.Icons.WAREHOUSE, size=18), ft.Text("Exportar bodega seleccionada")], spacing=8),
                            on_click=lambda _: self._exportar("bodega"),
                        ),
                    ],
                ),
            ], spacing=8)),
        ], vertical_alignment=ft.CrossAxisAlignment.CENTER)

        _header_col = ft.Column([_header, self._cont_preview], spacing=8)

        _body = ft.Row([
            ft.Container(
                width=190 if self._ver_categorias else 0,
                visible=self._ver_categorias,
                content=ft.Column([
                    ft.Container(content=ft.Text("FAMILIAS", size=10, weight="bold", color="grey500"),
                                 padding=ft.Padding(8, 6, 8, 4)),
                    ft.Container(content=self.col_arbol, expand=True,
                                 border_radius=10, bgcolor="white",
                                 border=ft.border.all(1, "grey100"), padding=4),
                ], expand=True),
                padding=ft.Padding(0, 0, 10, 0),
            ),
            ft.Container(
                expand=True,
                content=ft.Column([
                    ft.Container(
                        content=ft.Column([ft.Row([self.tabla], scroll=ft.ScrollMode.ALWAYS)],
                                         scroll=ft.ScrollMode.ALWAYS),
                        expand=True, bgcolor="white", border_radius=10,
                        border=ft.border.all(1, "grey100"), padding=0,
                        clip_behavior=ft.ClipBehavior.HARD_EDGE,
                    ),
                ], expand=True),
            ),
        ], expand=True, vertical_alignment=ft.CrossAxisAlignment.START, scroll=ft.ScrollMode.AUTO)

        self._main_content = sandwich(_header_col, _body, self.row_paginacion)

    def mostrar_snack(self, mensaje, tipo_o_color="success"):
        from utils.ui_helpers import mostrar_snack as _snack
        _snack(self.page_ref, mensaje, tipo_o_color)

    # ══════════════════════════════ Ciclo ══════════════════════════════
    def _poblar_desde_cache(self, app_cache):
        """Aplica datos del cache a los controles y hace content-swap. Sin update()."""
        bodegas = app_cache.inv_bodegas or []
        self.dd_bodega.options = [ft.dropdown.Option("0", "Global (Todas)")] + \
            [ft.dropdown.Option(str(b["id"]), _label_bodega(b)) for b in bodegas]
        self._arbol_data = app_cache.inv_arbol_data or {}
        self._sku_bodegas_map = app_cache.inv_sku_bodegas_map or {}
        self._build_arbol()
        prods = list(app_cache.inv_productos or [])
        self._productos_base = prods
        self._actualizar_tabla_ui(prods, {}, 0, actualizar=False)
        self.content = self._main_content
        self.datos_cargados = True
        self._cargando = False

    def inicializar(self):
        if self.datos_cargados or self._cargando:
            return
        self._cargando = True
        from core.utilidades import hilo as _hilo
        _hilo(self._cargar_desde_cache)

    def _cargar_desde_cache(self):
        """Espera el evento de cache y renderiza — solo se usa si el cache no estaba listo al navegar."""
        from core.cache import app_cache
        app_cache.inv_ready.wait(timeout=60)
        print(f"[INV] cache tardó — renderizando desde hilo de espera")
        self._poblar_desde_cache(app_cache)
        if self.page_ref:
            self.page_ref.update()

    def _recargar(self, e=None):
        from core.cache import app_cache
        app_cache.inv_ready.clear()
        self.datos_cargados = False
        self._cargando = True
        if self.page_ref:
            self.page_ref.update()
        self._cargar_todo()

    def _buscar_preview(self, e=None):
        texto = (self.in_buscar.value or "").strip()
        if not texto:
            self._cont_preview.visible = False
            if self.page_ref: self._cont_preview.update()
            return
        def _run():
            try:
                # Buscar por SKU exacto primero, luego por nombre
                res = get_sb().table("productos").select("sku, nombre, stock_global").eq("sku", texto).execute()
                if not res.data:
                    res = get_sb().table("productos").select("sku, nombre, stock_global").ilike("nombre", f"%{texto}%").execute()
                if res.data and len(res.data) == 1:
                    p = res.data[0]
                    stk = p.get("stock_global") or 0
                    color_stk = "red600" if stk == 0 else "orange600" if stk <= 3 else "green700"
                    self._txt_preview.value = f"📦  {p['nombre']}   ·   SKU: {p['sku']}   ·   Stock: {stk} uds"
                    self._txt_preview.color = "blue900"
                    self._cont_preview.bgcolor = "blue50"
                    self._cont_preview.border = ft.Border(left=ft.BorderSide(3, color_stk))
                    self._cont_preview.visible = True
                elif res.data and len(res.data) > 1:
                    self._cont_preview.visible = False
                else:
                    self._txt_preview.value = f"⚠️  No se encontró ningún producto con '{texto}'"
                    self._txt_preview.color = "red700"
                    self._cont_preview.bgcolor = "red50"
                    self._cont_preview.border = ft.Border(left=ft.BorderSide(3, "red400"))
                    self._cont_preview.visible = True
                if self.page_ref: self._cont_preview.update()
            except Exception:
                pass
        hilo(_run)

    def _limpiar(self, e=None):
        self.in_buscar.value = ""
        self._cont_preview.visible = False
        self._fam_actual = "Todas"
        self._subfam_actual = "Todas"
        self.txt_filtro.value = "Mostrando: Todas"
        self._build_arbol()
        if self.page_ref:
            self.in_buscar.update()
            self._cont_preview.update()
            self.txt_filtro.update()
            self.col_arbol.update()
        self._cargar_tabla()

    def _cambiar_bodega(self, e):
        self._cargar_tabla()

    # ══════════════════════════════ Carga ══════════════════════════════
    def _cargar_todo(self):
        """Re-obtiene datos desde la BD, actualiza cache y hace content-swap."""
        def _run():
            try:
                from core.cache import app_cache
                sb = get_sb()

                # 1. Bodegas
                bodegas = sb.table("bodegas").select("id, nombre, areas(nombre)").execute().data or []
                self.dd_bodega.options = [ft.dropdown.Option("0", "Global (Todas)")] + \
                    [ft.dropdown.Option(str(b["id"]), _label_bodega(b)) for b in bodegas]

                # 2. Árbol familias
                fams = sb.table("familias").select("id, nombre").order("nombre").execute().data or []
                subs = sb.table("subfamilias").select("nombre, familia_id").execute().data or []
                arbol = {f["nombre"]: [] for f in fams}
                fam_id_map = {f["id"]: f["nombre"] for f in fams}
                for s in subs:
                    fn = fam_id_map.get(s["familia_id"])
                    if fn in arbol:
                        arbol[fn].append(s["nombre"])
                self._arbol_data = arbol
                self._build_arbol()

                # 3. Productos + tabla (sincrónico, sin update() intermedios)
                self._cargar_productos_base(sb)
                bodega_id = int(self.dd_bodega.value or "0")
                stock_map = {}
                if bodega_id:
                    rows_bp = fetch_all("bodega_productos", "sku, cantidad", bodega_id=bodega_id)
                    stock_map = {r["sku"]: r["cantidad"] for r in rows_bp}
                prods = self._productos_base
                if bodega_id:
                    prods = [p for p in prods if p["sku"] in stock_map]
                self._actualizar_tabla_ui(prods, stock_map, bodega_id, actualizar=False)

                # 4. Construir map bodega/area por SKU
                bp_raw = sb.table("bodega_productos") \
                    .select("sku, cantidad, bodegas(nombre, areas(nombre))").execute().data or []
                sku_bod_map: dict = {}
                for r in bp_raw:
                    if (r.get("cantidad") or 0) <= 0:
                        continue
                    sku = r["sku"]
                    bod = r.get("bodegas") or {}
                    area = (bod.get("areas") or {}).get("nombre") or ""
                    sku_bod_map.setdefault(sku, []).append(
                        {"bodega": bod.get("nombre") or "", "area": area}
                    )
                self._sku_bodegas_map = sku_bod_map

                # 5. Actualizar cache con datos frescos
                app_cache.inv_bodegas = bodegas
                app_cache.inv_arbol_data = arbol
                app_cache.inv_productos = list(self._productos_base)
                app_cache.inv_sku_bodegas_map = sku_bod_map
                app_cache.inv_ready.set()

                # 5. Content-swap — envía todo el árbol con datos al cliente de una vez
                self.content = self._main_content
                self.datos_cargados = True
                self._cargando = False
                if self.page_ref:
                    self.page_ref.update()

            except Exception as ex:
                self._cargando = False
                self.mostrar_snack(f"❌ Error al cargar inventario: {ex}", "red700")
        hilo(_run)

    def _cargar_productos_base(self, sb=None):
        """Descarga todos los productos con nombre de familia resuelto."""
        if sb is None:
            sb = get_sb()
        # JOIN con familias y subfamilias via Supabase (nuevo schema)
        raw = fetch_all("productos",
                        "sku, nombre, costo_neto, precio_venta, stock_global, "
                        "familia_id, subfamilia_id, "
                        "familias(nombre), subfamilias(nombre)")
        prods = []
        for p in raw:
            fam_obj = p.get("familias") or {}
            sub_obj = p.get("subfamilias") or {}
            prods.append({
                "sku":         p.get("sku", ""),
                "nombre":      p.get("nombre", ""),
                "familia":     fam_obj.get("nombre") if isinstance(fam_obj, dict) else "",
                "subfamilia":  sub_obj.get("nombre") if isinstance(sub_obj, dict) else "",
                "familia_id":  p.get("familia_id"),
                "subfamilia_id": p.get("subfamilia_id"),
                "stock_actual": p.get("stock_global", 0) or 0,
                "costo_neto":  p.get("costo_neto", 0) or 0,
                "precio_venta": p.get("precio_venta", 0) or 0,
            })
        self._productos_base = prods

    def _cargar_tabla(self):
        def _run():
            try:
                sb = get_sb()
                bodega_id = int(self.dd_bodega.value or "0")
                sku_filtro = (self.in_buscar.value or "").strip().lower()

                # Si hay bodega seleccionada, obtener stock específico
                stock_map = {}
                if bodega_id and bodega_id != 0:
                    rows = fetch_all("bodega_productos", "sku, cantidad", bodega_id=bodega_id)
                    stock_map = {r["sku"]: r["cantidad"] for r in rows}

                # Recargar productos base si faltan SKUs (ej: importación reciente)
                if not self._productos_base:
                    self._cargar_productos_base(sb)
                elif stock_map:
                    prods_skus = {p["sku"] for p in self._productos_base}
                    if any(sku not in prods_skus for sku in stock_map):
                        self._cargar_productos_base(sb)

                # Filtrar
                prods = self._productos_base
                if sku_filtro:
                    prods = [p for p in prods if sku_filtro in p["sku"].lower() or sku_filtro in p["nombre"].lower()]
                else:
                    if self._fam_actual != "Todas":
                        prods = [p for p in prods if (p.get("familia") or "") == self._fam_actual]
                    if self._subfam_actual != "Todas":
                        prods = [p for p in prods if (p.get("subfamilia") or "") == self._subfam_actual]
                    if bodega_id != 0:
                        prods = [p for p in prods if p["sku"] in stock_map]

                self._actualizar_tabla_ui(prods, stock_map, bodega_id)
            except Exception as ex:
                self.mostrar_snack(f"❌ Error al filtrar tabla: {ex}", "red700")
        hilo(_run)

    def _cargar_tabla_con_productos(self):
        """Llama _cargar_tabla usando los productos ya en memoria."""
        self._cargar_tabla()

    def _actualizar_tabla_ui(self, prods, stock_map, bodega_id, actualizar=True):
        self._prods_filtrados = prods
        self._stock_map_actual = stock_map
        self._bodega_id_actual = bodega_id
        self._pagina = 0
        self._renderizar_pagina(actualizar=actualizar)

    def _saltar_pagina(self, e=None):
        try:
            destino = int(self.in_ir_pagina.value or "1") - 1
        except:
            return
        self.in_ir_pagina.value = ""
        if self.page_ref: self.in_ir_pagina.update()
        self._ir_pagina(destino)

    def _ir_pagina(self, nueva_pagina):
        total = len(self._prods_filtrados)
        max_pag = max(0, (total - 1) // self._por_pagina)
        self._pagina = max(0, min(nueva_pagina, max_pag))
        self._renderizar_pagina()

    def _renderizar_pagina(self, actualizar=True):
        prods = self._prods_filtrados
        stock_map = getattr(self, "_stock_map_actual", {})
        bodega_id = getattr(self, "_bodega_id_actual", 0)
        total = len(prods)
        inicio = self._pagina * self._por_pagina
        fin = min(inicio + self._por_pagina, total)
        pagina_prods = prods[inicio:fin]

        nuevas_filas = []
        for p in pagina_prods:
            stk = stock_map.get(p["sku"], p["stock_actual"]) if (bodega_id and bodega_id != 0) else p["stock_actual"]
            stk = stk or 0
            celdas = [ft.DataCell(ft.Text(p["sku"], weight="bold", size=12, color="blue900"))]
            if self._ver_categorias:
                celdas += [
                    ft.DataCell(ft.Text(p.get("familia") or "—", size=12, color="grey700")),
                    ft.DataCell(ft.Text(p.get("subfamilia") or "—", size=12, color="blue700")),
                ]
            # Bodega / Área
            sku_bods = self._sku_bodegas_map.get(p["sku"], [])
            if bodega_id and bodega_id != 0:
                match = next((b for b in sku_bods if b["bodega"]), None)
                if match:
                    area_txt = match["area"]
                    bod_lbl = f"[{area_txt}] {match['bodega']}" if area_txt else match["bodega"]
                else:
                    bod_lbl = "—"
            else:
                if not sku_bods:
                    bod_lbl = "—"
                elif len(sku_bods) == 1:
                    area_txt = sku_bods[0]["area"]
                    bod_lbl = f"[{area_txt}] {sku_bods[0]['bodega']}" if area_txt else sku_bods[0]["bodega"]
                else:
                    partes = []
                    for b in sku_bods:
                        partes.append(f"[{b['area']}] {b['bodega']}" if b["area"] else b["bodega"])
                    bod_lbl = " · ".join(partes)

            celdas += [
                ft.DataCell(ft.Text(p["nombre"], size=12)),
                ft.DataCell(ft.Text(bod_lbl, size=11, color="indigo700")),
                ft.DataCell(badge_stock(stk)),
                ft.DataCell(ft.Text(clp(p['precio_venta']), size=12, color="grey800")),
            ]
            if estado.puede_ver("costo"):
                celdas.append(ft.DataCell(ft.Text(clp(p['costo_neto']), size=12, color="red700")))
            nuevas_filas.append(ft.DataRow(cells=celdas))

        self.tabla.rows = nuevas_filas
        total_pags = max(1, -(-total // self._por_pagina))  # ceil division
        self.lbl_pagina.value = f"Página {self._pagina + 1} de {total_pags}  ({total} productos)"
        self.btn_prev.disabled = self._pagina == 0
        self.btn_next.disabled = self._pagina >= total_pags - 1
        if actualizar and self.page_ref:
            self.tabla.update()
            self.lbl_pagina.update()
            self.btn_prev.update()
            self.btn_next.update()

    # ══════════════════════════════ Árbol ══════════════════════════════
    def _build_arbol(self):
        self.col_arbol.controls.clear()
        self.col_arbol.controls.append(ft.Container(
            content=ft.Text("📦 Todas las familias", size=12, weight="bold", color="blue900"),
            padding=ft.Padding(10, 6, 10, 6), border_radius=6,
            bgcolor="blue100" if self._fam_actual == "Todas" else "transparent",
            ink=True, on_click=lambda _: self._sel_arbol("Todas", "Todas"),
        ))
        for fam, subs in sorted(self._arbol_data.items()):
            expandido = fam in self._arbol_expandido
            fam_activa = (self._fam_actual == fam)

            def _click_fam(e, f=fam):
                self._arbol_expandido.add(f)
                self._sel_arbol(f, "Todas")

            def _toggle(e, f=fam):
                if f in self._arbol_expandido: self._arbol_expandido.discard(f)
                else: self._arbol_expandido.add(f)
                self._build_arbol()
                if self.page_ref: self.col_arbol.update()

            fila_fam = ft.Container(
                content=ft.Row([ft.Text(fam, size=12,
                    weight="bold" if fam_activa else "normal",
                    color="blue900" if fam_activa else "grey800", expand=True)], spacing=2),
                padding=ft.Padding(6, 5, 6, 5), border_radius=6,
                bgcolor="blue50" if (fam_activa and self._subfam_actual == "Todas") else "transparent",
                ink=True, on_click=_click_fam,
            )
            toggle_btn = ft.Container(
                content=ft.Icon(ft.Icons.ARROW_DROP_DOWN if expandido else ft.Icons.ARROW_RIGHT, size=20, color="blue400"),
                on_click=_toggle, width=26, padding=ft.Padding(0, 4, 0, 4),
            )
            self.col_arbol.controls.append(ft.Row([toggle_btn, ft.Container(content=fila_fam, expand=True)], spacing=0))

            if expandido:
                for sub in sorted(subs):
                    sub_activa = fam_activa and self._subfam_actual == sub
                    self.col_arbol.controls.append(ft.Container(
                        content=ft.Row([
                            ft.Container(width=26),
                            ft.Icon(ft.Icons.SUBDIRECTORY_ARROW_RIGHT, size=14, color="grey400"),
                            ft.Text(sub, size=11,
                                    color="blue800" if sub_activa else "grey700",
                                    weight="bold" if sub_activa else "normal"),
                        ], spacing=4),
                        padding=ft.Padding(4, 4, 4, 4), border_radius=5,
                        bgcolor="blue100" if sub_activa else "transparent",
                        ink=True,
                        on_click=lambda e, f=fam, s=sub: self._sel_arbol(f, s),
                    ))

    def _sel_arbol(self, familia, subfamilia):
        self._fam_actual = familia or "Todas"
        self._subfam_actual = subfamilia or "Todas"
        label = "Todas las familias" if self._fam_actual == "Todas" else (
            f"{self._fam_actual} › {self._subfam_actual}" if self._subfam_actual != "Todas" else self._fam_actual
        )
        self.txt_filtro.value = f"Filtrando: {label}"
        self._build_arbol()
        if self.page_ref:
            self.txt_filtro.update()
            self.col_arbol.update()
        self._cargar_tabla()

    # ══════════════════════════════ Exportar ══════════════════════════════
    def _exportar(self, modo: str):
        def _run():
            import openpyxl, io, base64 as _b64
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime

            sb = get_sb()

            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill("solid", fgColor="1565C0")
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin      = Side(style="thin", color="CCCCCC")
            borde     = Border(left=thin, right=thin, top=thin, bottom=thin)

            def estilizar_hoja(ws, encabezados):
                ws.row_dimensions[1].height = 22
                for ci, h in enumerate(encabezados, 1):
                    c = ws.cell(1, ci, h)
                    c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, borde

            def autoajustar(ws):
                for col in ws.columns:
                    max_w = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 50)

            def aplicar_formatos_clp(ws, cols_precio: list):
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.number_format = "@"
                for col_idx in cols_precio:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.number_format = "#,##0"

            try:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)  # quitar hoja por defecto

                ver_costo = estado.puede_ver("costo")

                # ════════════ GLOBAL ════════════
                if modo == "global":
                    prods = self._productos_base
                    ws = wb.create_sheet("Inventario Global")
                    cols = ["SKU", "Nombre", "Familia", "Subfamilia", "Stock Global", "Precio Venta ($)"]
                    if ver_costo:
                        cols += ["Costo Neto ($)", "Margen ($)"]
                    estilizar_hoja(ws, cols)
                    for p in prods:
                        fila = [
                            p["sku"], p["nombre"],
                            p.get("familia") or "", p.get("subfamilia") or "",
                            p.get("stock_actual", 0),
                            p.get("precio_venta", 0),
                        ]
                        if ver_costo:
                            costo = p.get("costo_neto", 0)
                            fila += [costo, (p.get("precio_venta", 0) or 0) - (costo or 0)]
                        ws.append(fila)
                    autoajustar(ws)
                    aplicar_formatos_clp(ws, [6, 7, 8] if ver_costo else [6])

                # ════════════ BODEGA ════════════
                elif modo == "bodega":
                    bodega_id = int(self.dd_bodega.value or "0")
                    if not bodega_id:
                        # Exportar todas las bodegas, una hoja por bodega
                        bodegas = sb.table("bodegas").select("id, nombre").order("id").execute().data or []
                    else:
                        nom = next((o.text for o in self.dd_bodega.options if o.key == str(bodega_id)), f"Bodega {bodega_id}")
                        bodegas = [{"id": bodega_id, "nombre": nom}]

                    prods_base = {p["sku"]: p for p in self._productos_base}

                    for bod in bodegas:
                        filas_bp = fetch_all("bodega_productos", "sku, cantidad", bodega_id=bod["id"])
                        ws = wb.create_sheet(bod["nombre"][:31])
                        cols = ["SKU", "Nombre", "Familia", "Subfamilia", "Stock en Bodega", "Precio Venta ($)"]
                        if ver_costo:
                            cols += ["Costo Neto ($)", "Valor Total Costo ($)"]
                        estilizar_hoja(ws, cols)
                        for r in filas_bp:
                            p = prods_base.get(r["sku"], {})
                            cant = r.get("cantidad", 0) or 0
                            fila = [
                                r["sku"],
                                p.get("nombre", ""),
                                p.get("familia", "") or "",
                                p.get("subfamilia", "") or "",
                                cant,
                                p.get("precio_venta", 0),
                            ]
                            if ver_costo:
                                costo = p.get("costo_neto", 0) or 0
                                fila += [costo, cant * costo]
                            ws.append(fila)
                        autoajustar(ws)
                        aplicar_formatos_clp(ws, [6, 7, 8] if ver_costo else [6])

                # ════════════ FURGONES ════════════
                elif modo == "furgones":
                    furgones = sb.table("furgones").select("id, nombre").order("nombre").execute().data or []
                    prods_base = {p["sku"]: p for p in self._productos_base}

                    for furg in furgones:
                        filas_fp = sb.table("furgon_productos").select("sku, cantidad") \
                                     .eq("furgon_id", furg["id"]).execute().data or []
                        ws = wb.create_sheet(furg["nombre"][:31])
                        cols = ["SKU", "Nombre", "Familia", "Subfamilia", "Cantidad", "Precio Venta ($)"]
                        if ver_costo:
                            cols += ["Costo Neto ($)", "Valor Total Costo ($)"]
                        estilizar_hoja(ws, cols)
                        for r in filas_fp:
                            p = prods_base.get(r["sku"], {})
                            cant = r.get("cantidad", 0) or 0
                            fila = [
                                r["sku"],
                                p.get("nombre", ""),
                                p.get("familia", "") or "",
                                p.get("subfamilia", "") or "",
                                cant,
                                p.get("precio_venta", 0),
                            ]
                            if ver_costo:
                                costo = p.get("costo_neto", 0) or 0
                                fila += [costo, cant * costo]
                            ws.append(fila)
                        autoajustar(ws)
                        aplicar_formatos_clp(ws, [6, 7, 8] if ver_costo else [6])

                if not wb.sheetnames:
                    wb.create_sheet("Sin datos")

                _buf = io.BytesIO()
                wb.save(_buf)
                import subprocess as _sp, datetime as _dt
                _fname = f"/tmp/Inventario_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                open(_fname, 'wb').write(_buf.getvalue())
                _sp.Popen(['open', _fname])
                from utils.ui_helpers import mostrar_snack
                mostrar_snack(self.page_ref, "✅ Excel generado correctamente.", "success")

            except Exception as ex:
                from utils.ui_helpers import mostrar_snack
                mostrar_snack(self.page_ref, f"Error al exportar: {ex}", "error")

        hilo(_run)