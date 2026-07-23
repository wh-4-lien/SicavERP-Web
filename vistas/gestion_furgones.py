# vistas/gestion_furgones.py
import flet as ft
import datetime
from core.database import get_sb, cache_invalidate, fetch_herramientas_furgo
from core.estado import estado
from core.utilidades import hilo, registrar_auditoria
from core.widgets import BuscadorProducto


def _label_bodega(b: dict) -> str:
    area = (b.get("areas") or {}).get("nombre") or ""
    return f"[{area}] {b['nombre']}" if area else b["nombre"]

class VistaGestionFurgones(ft.Container):
    def __init__(self, page: ft.Page):
        super().__init__()
        self.page_ref = page
        self.expand = True
        self._opciones_bodegas = []
        self.construir_ui()

    def mostrar_snack(self, mensaje, tipo_o_color="success"):
        from utils.ui_helpers import mostrar_snack as _snack
        _snack(self.page_ref, mensaje, tipo_o_color)

    def mostrar_dialogo_importacion(self, exito: bool, titulo: str, mensaje: str, detalle: str = ""):
        # Definición del diálogo
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.CHECK_CIRCLE if exito else ft.Icons.ERROR, 
                        color="green" if exito else "red", size=28),
                ft.Text(titulo, weight="bold", size=16),
            ]),
            content=ft.Column([
                ft.Text(mensaje, size=13),
                ft.Text(detalle, size=11, italic=True) if detalle else ft.Container(),
            ], tight=True),
        )

        def _cerrar(e=None):
            self.page_ref.pop_dialog()

        dlg.actions = [ft.TextButton("Cerrar", on_click=_cerrar)]

        self.page_ref.show_dialog(dlg)

    def construir_ui(self):
        self.in_furgo_nombre_nuevo = ft.TextField(label="Nombre del furgón", width=220, height=50, hint_text="Ej: Furgón 01")
        self.in_furgo_desc_nuevo = ft.TextField(label="Descripción / Patente", width=280, height=50, hint_text="Ej: Ford Transit - ABC123")
        self.sel_furgo_tecnico = ft.Dropdown(label="Asignar a técnico", width=280, options=[])
        self.txt_gf_estado = ft.Text("", size=12, color="grey500", italic=True)
        
        self.col_lista_furgones = ft.Column([], spacing=6)
        self.txt_import_furgo = ft.Text("", size=12, color="grey600")
        self.col_furgo_gestion_detalle = ft.Column([], spacing=10, scroll=ft.ScrollMode.AUTO)

        self.cont_crear = ft.Container(
            content=ft.Column([
                ft.Text("➕ Crear nuevo furgón", weight="bold", size=15, color="blue900"),
                ft.Divider(height=8, color="transparent"),
                ft.Row([self.in_furgo_nombre_nuevo, self.in_furgo_desc_nuevo], spacing=12),
                ft.Row([
                    self.sel_furgo_tecnico,
                    ft.ElevatedButton("CREAR FURGÓN", icon=ft.Icons.ADD, on_click=self.crear_furgon, bgcolor="blue700", color="white", height=50, tooltip="Crear nuevo furgón"),
                ], spacing=12),
                self.txt_gf_estado,
            ]),
            padding=20, bgcolor="white", border_radius=12, border=ft.border.all(1, "grey200"),
            visible=estado.puede_hacer("gestion_furgones", "crear_furgon"),
        )

        self.cont_lista = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Text("🚐 Furgones existentes", weight="bold", size=15, color="blue900", expand=True),
                    ft.ElevatedButton("EXPORTAR PDF GLOBAL", icon=ft.Icons.PICTURE_AS_PDF, bgcolor="red700", color="white", height=38, on_click=self.exportar_pdf_inventario_furgon),
                    ft.ElevatedButton("🔄 Actualizar", height=38, on_click=lambda _: self.cargar_datos()),
                ]),
                ft.Divider(height=8, color="transparent"),
                self.col_lista_furgones,
            ]),
            padding=20, bgcolor="white", border_radius=12, border=ft.border.all(1, "grey200"),
            visible=estado.es_admin,
        )

        self.cont_gestion = ft.Container(
            content=ft.Column([
                ft.Text("📥 Importar stock / Ajustar furgones", weight="bold", size=14, color="grey800"),
                ft.Divider(height=6, color="transparent"),
                self.col_furgo_gestion_detalle,
            ]),
            padding=20, bgcolor="white", border_radius=10, border=ft.border.all(1, "grey200"),
        )

        from utils.ui_helpers import sandwich

        _header = ft.Row([
            ft.Text("Gestión de Furgones", size=24, weight="bold", color="blue900", expand=True),
            ft.IconButton(ft.Icons.REFRESH, tooltip="Actualizar lista de furgones", on_click=lambda _: self.cargar_datos()),
        ])

        _body = ft.Column([
            self.cont_crear,
            ft.Divider(height=14, color="transparent"),
            self.cont_lista,
            ft.Divider(height=14, color="transparent"),
            self.cont_gestion,
        ], scroll=ft.ScrollMode.AUTO, expand=True)

        self.content = sandwich(_header, _body)

    def pre_cargar_si_cache_listo(self):
        from core.cache import app_cache
        if getattr(self, '_combos_listos', False):
            return
        if app_cache.inv_ready.is_set():
            bodegas = app_cache.inv_bodegas or []
            self._opciones_bodegas = [ft.dropdown.Option(key=str(b["id"]), text=b["nombre"]) for b in bodegas]
        if app_cache.gfur_ready.is_set():
            usuarios = app_cache.gfur_usuarios or []
            self.sel_furgo_tecnico.options.clear()
            self.sel_furgo_tecnico.options.append(ft.dropdown.Option("", "— Sin asignar —"))
            for u in usuarios:
                rol_nom = (u.get("roles") or {}).get("nombre", "")
                self.sel_furgo_tecnico.options.append(ft.dropdown.Option(u["usuario"], f"{u['nombre']} ({rol_nom})"))
            self._combos_listos = True
            print(f"[GESTION_FURGONES] combos pre-cargados ({len(usuarios)} técnicos)")

    def inicializar(self):
        if getattr(self, '_datos_cargados', False):
            return
        self.cargar_datos()

    def cargar_datos(self):
        self.txt_gf_estado.value = ""
        es_admin = estado.es_admin
        usuario_login = estado.usuario_actual.get("usuario", "")

        def _run():
            try:
                from core.cache import app_cache

                # Bodegas: reusar cache si disponible
                if app_cache.inv_ready.is_set() and app_cache.inv_bodegas:
                    self._opciones_bodegas = [ft.dropdown.Option(key=str(b["id"]), text=_label_bodega(b)) for b in app_cache.inv_bodegas]
                else:
                    res_b = get_sb().table("bodegas").select("id, nombre, areas(nombre)").order("id").execute()
                    self._opciones_bodegas = [ft.dropdown.Option(key=str(b["id"]), text=_label_bodega(b)) for b in (res_b.data or [])]

                # Usuarios y furgones: reusar cache si disponible
                if app_cache.gfur_ready.is_set() and app_cache.gfur_usuarios is not None:
                    usuarios_data = app_cache.gfur_usuarios
                    furgones_data = list(app_cache.gfur_furgones or [])
                    fp_counts = app_cache.gfur_fp_counts or {}
                else:
                    res_u = get_sb().table("usuarios").select("usuario, nombre, roles(nombre)").execute()
                    res_f = get_sb().table("furgones").select("*").order("id").execute()
                    usuarios_data = res_u.data or []
                    furgones_data = res_f.data or []
                    fp_counts = {}

                # Técnico: filtrar solo su furgón asignado (salvo que tenga permiso ver_todos)
                if not es_admin and not estado.puede_hacer("gestion_furgones", "ver_todos"):
                    furgones_data = [f for f in furgones_data if f.get("tecnico_usuario") == usuario_login]

                if es_admin:
                    self.sel_furgo_tecnico.options.clear()
                    self.sel_furgo_tecnico.options.append(ft.dropdown.Option("", "— Sin asignar —"))
                    for u in usuarios_data:
                        rol_nom = (u.get("roles") or {}).get("nombre", "")
                        self.sel_furgo_tecnico.options.append(ft.dropdown.Option(u["usuario"], f"{u['nombre']} ({rol_nom})"))

                nuevos_controles_lista = []
                effective_fp_counts = dict(fp_counts) if fp_counts else {}

                if not furgones_data:
                    msg = "No hay furgones creados aún." if es_admin else "No tienes un furgón asignado."
                    nuevos_controles_lista.append(ft.Text(msg, size=12, color="grey400", italic=True))
                    self.col_furgo_gestion_detalle.controls = [ft.Text(msg, color="grey400", italic=True, size=13)]
                else:
                    for f in furgones_data:
                        fid = f["id"]
                        fname = f["nombre"]

                        if fid in effective_fp_counts:
                            n_prod = effective_fp_counts[fid]
                        else:
                            res_p = get_sb().table("furgon_productos").select("id").eq("furgon_id", fid).execute()
                            n_prod = len(res_p.data or [])
                            effective_fp_counts[fid] = n_prod

                        puede_eliminar_f = estado.puede_hacer("gestion_furgones", "eliminar_furgon")
                        if es_admin:
                            sel_reasignar = ft.Dropdown(
                                width=240, height=45, value=f.get("tecnico_usuario") or "",
                                options=[ft.dropdown.Option("", "— Sin asignar —")] + [
                                    ft.dropdown.Option(u["usuario"], f"{u['nombre']} ({(u.get('roles') or {}).get('nombre', '')})") for u in usuarios_data
                                ],
                            )
                            nuevos_controles_lista.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Icon(ft.Icons.LOCAL_SHIPPING, color="blue700", size=22),
                                            ft.Column([
                                                ft.Text(fname, weight="bold", size=14, color="grey900"),
                                                ft.Text(f.get("descripcion") or "", size=11, color="grey500"),
                                            ], expand=True, spacing=1),
                                            ft.Container(content=ft.Text(f"{n_prod} productos", size=11, color="white"), bgcolor="blue600", border_radius=20, padding=ft.Padding(10, 3, 10, 3)),
                                            ft.IconButton(ft.Icons.DELETE_OUTLINE, icon_color="red400", tooltip="Eliminar furgón", on_click=self.confirmar_eliminacion_furgo(fid, fname), visible=puede_eliminar_f),
                                        ]),
                                        ft.Row([
                                            ft.Text("Técnico asignado:", size=12, color="grey600"),
                                            sel_reasignar,
                                            ft.ElevatedButton("GUARDAR", height=40, on_click=self.hacer_reasignar(fid, sel_reasignar), bgcolor="teal600", color="white"),
                                        ], spacing=10),
                                    ], spacing=8),
                                    padding=16, bgcolor="white", border_radius=10, border=ft.border.all(1, "grey200"),
                                )
                            )

                    # Cache para re-uso en tarjetas
                    self._furgones_data_cache = list(furgones_data)
                    self._fp_counts_cache = effective_fp_counts

                    cards = [self._construir_card_furgon_gestion(f, effective_fp_counts.get(f["id"], 0)) for f in furgones_data]
                    cards.append(self.txt_import_furgo)
                    self.col_furgo_gestion_detalle.controls = cards

                self.col_lista_furgones.controls = nuevos_controles_lista
                self.cont_lista.visible = es_admin
                self._datos_cargados = True

                if self.page_ref:
                    self.col_lista_furgones.update()
                    self.col_furgo_gestion_detalle.update()
                    self.cont_lista.update()
            except Exception as ex:
                self.mostrar_snack(f"❌ Error al cargar furgones: {ex}", "red700")
        hilo(_run)

    # --- NUEVA FUNCIÓN PARA VER INVENTARIO COMO ADMIN ---
    def ver_inventario_furgo(self, fid, fname):
        def _abrir(e):
            def _run():
                try:
                    res = get_sb().table("furgon_productos").select("cantidad, sku, productos(nombre)").eq("furgon_id", fid).execute()
                    
                    filas = []
                    for p in (res.data or []):
                        sku = p.get("sku", "")
                        cant = p.get("cantidad", 0)
                        nom = p.get("productos", {}).get("nombre", "") if isinstance(p.get("productos"), dict) else sku
                        
                        color_cant = "red700" if cant == 0 else "orange700" if cant <= 2 else "green700"
                        lbl_cant = ft.Container(
                            content=ft.Text(str(cant), weight="bold", color="white", size=12),
                            bgcolor=color_cant, border_radius=12, padding=ft.Padding(10, 4, 10, 4)
                        )
                        
                        filas.append(ft.DataRow(cells=[
                            ft.DataCell(ft.Text(sku, weight="bold", color="blue900", size=13)),
                            ft.DataCell(ft.Text(nom, size=13, color="grey800")),
                            ft.DataCell(lbl_cant)
                        ]))
                    
                    filas.sort(key=lambda r: r.cells[1].content.value)
                    
                    if not filas:
                        contenido = ft.Text("El furgón no tiene productos asignados actualmente.", size=14, italic=True, color="grey600")
                    else:
                        tabla = ft.DataTable(
                            columns=[
                                ft.DataColumn(ft.Text("SKU", weight="bold")),
                                ft.DataColumn(ft.Text("Producto", weight="bold")),
                                ft.DataColumn(ft.Text("Cantidad", weight="bold"), numeric=True),
                            ],
                            rows=filas,
                            border=ft.border.all(1, "grey200"),
                            border_radius=8,
                            heading_row_color=ft.Colors.with_opacity(0.04, "blue")
                        )
                        contenido = ft.Column([ft.Row([tabla], scroll=ft.ScrollMode.ALWAYS)], scroll=ft.ScrollMode.ALWAYS)

                    dlg_inv = ft.AlertDialog(
                        modal=True,
                        title=ft.Row([ft.Icon(ft.Icons.INVENTORY, color="blue900"), ft.Text(f"Inventario: {fname}", weight="bold")], spacing=10),
                        content=ft.Container(content=contenido, width=600, height=400, padding=10),
                        actions_alignment=ft.MainAxisAlignment.END,
                    )

                    def _cerrar_inv(e_cerrar):
                        if hasattr(self.page_ref, "close"): self.page_ref.close(dlg_inv)
                        else:
                            dlg_inv.open = False
                            try: dlg_inv.update()
                            except: pass
                            if dlg_inv in self.page_ref.overlay: self.page_ref.overlay.remove(dlg_inv)
                            self.page_ref.update()

                    dlg_inv.actions = [ft.TextButton("Cerrar", on_click=_cerrar_inv)]

                    if hasattr(self.page_ref, "open"): self.page_ref.open(dlg_inv)
                    else:
                        self.page_ref.overlay.append(dlg_inv)
                        dlg_inv.open = True
                        self.page_ref.update()

                except Exception as ex:
                    self.mostrar_snack(f"❌ Error al cargar el inventario del furgón: {ex}", "red700")
            hilo(_run)
        return _abrir
    # ----------------------------------------------------

    def _construir_card_furgon_gestion(self, f, n_prod):
        fid = f["id"]
        fname = f["nombre"]
        tec = f.get("tecnico_usuario") or "Sin asignar"

        puede_importar = estado.puede_hacer("gestion_furgones", "importar")
        puede_ajustar = estado.puede_hacer("gestion_furgones", "ajustar_stock")

        buscador_adj = BuscadorProducto(label="SKU o Nombre", width=200, prefix_icon=ft.Icons.QR_CODE_SCANNER)
        in_cant_adj = ft.TextField(label="Cant.", value="1", width=90, height=50, text_align="center")
        dd_bodega_adj = ft.Dropdown(label="Bodega origen", width=220, height=50, options=self._opciones_bodegas)
        txt_adj_info = ft.Text("", size=11, color="grey600", italic=True)
        cont_adj_info = ft.Container(
            content=txt_adj_info, visible=False,
            padding=ft.Padding(10, 6, 10, 6), border_radius=6,
            bgcolor="blue50", border=ft.Border(left=ft.BorderSide(3, "blue400")),
        )
        buscador_adj._on_seleccionar = self.hacer_buscar_adj(buscador_adj, txt_adj_info, cont_adj_info, fid)

        controles_body = []
        if puede_importar:
            controles_body.append(ft.Row([
                ft.ElevatedButton("📥 PLANTILLA", icon=ft.Icons.DOWNLOAD, on_click=self.descargar_plantilla_furgo(fname), bgcolor="teal700", color="white", height=42),
                ft.ElevatedButton("🚀 IMPORTAR EXCEL", icon=ft.Icons.UPLOAD_FILE, on_click=self.importar_stock_furgo(fid, fname), bgcolor="blue700", color="white", height=42),
            ], spacing=10))

        if puede_ajustar:
            controles_body += [
                ft.Text("⚖️ Ajuste de stock", size=12, weight="bold", color="grey700"),
                ft.Row([
                    buscador_adj, in_cant_adj, dd_bodega_adj,
                    ft.ElevatedButton("SUMAR", icon=ft.Icons.ADD, on_click=self.hacer_ajuste_furgo(buscador_adj, in_cant_adj, dd_bodega_adj, txt_adj_info, cont_adj_info, fid, fname, 1), bgcolor="green700", color="white", height=50),
                    ft.ElevatedButton("RESTAR", icon=ft.Icons.REMOVE, on_click=self.hacer_ajuste_furgo(buscador_adj, in_cant_adj, dd_bodega_adj, txt_adj_info, cont_adj_info, fid, fname, -1), bgcolor="orange700", color="white", height=50),
                ], spacing=8),
                cont_adj_info,
            ]

        controles_body.append(self.construir_bloque_herramientas(fid, fname))

        body_col = ft.Column(controles_body, visible=False, spacing=12)
        chevron = ft.Icon(ft.Icons.KEYBOARD_ARROW_DOWN, color="grey500", size=22)

        def _toggle(e):
            body_col.visible = not body_col.visible
            chevron.name = ft.Icons.KEYBOARD_ARROW_UP if body_col.visible else ft.Icons.KEYBOARD_ARROW_DOWN
            body_col.update()
            chevron.update()

        header = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.LOCAL_SHIPPING, color="blue700", size=22),
                ft.Column([
                    ft.Text(fname, weight="bold", size=14, color="grey900"),
                    ft.Text(f.get("descripcion") or "", size=11, color="grey500"),
                    ft.Text(f"Técnico: {tec}", size=11, color="grey500"),
                ], expand=True, spacing=1),
                ft.Container(
                    content=ft.Text(f"{n_prod} productos", size=11, color="white"),
                    bgcolor="blue600", border_radius=20, padding=ft.Padding(10, 3, 10, 3),
                ),
                chevron,
            ]),
            padding=16,
            on_click=_toggle,
            ink=True,
            border_radius=ft.BorderRadius(10, 10, 0, 0),
        )

        body_wrap = ft.Container(
            content=body_col,
            padding=ft.Padding(16, 0, 16, 16),
            bgcolor="grey50",
            border_radius=ft.BorderRadius(0, 0, 10, 10),
        )

        return ft.Container(
            content=ft.Column([header, body_wrap], spacing=0),
            bgcolor="white",
            border_radius=10,
            border=ft.border.all(1, "grey200"),
        )

    def crear_furgon(self, e):
        nombre = self.in_furgo_nombre_nuevo.value.strip()
        if not nombre:
            self.mostrar_snack("⚠️ Ingresa un nombre para el furgón", "orange")
            return
        desc  = self.in_furgo_desc_nuevo.value.strip()
        tecnico = self.sel_furgo_tecnico.value or None
        def _run():
            try:
                get_sb().table("furgones").insert({
                    "nombre": nombre,
                    "descripcion": desc,
                    "tecnico_usuario": tecnico,
                }).execute()
                cache_invalidate("furgones_totales")
                registrar_auditoria(estado.usuario_actual["nombre"], "CREAR FURGÓN", f"Furgón: {nombre}")
                self.mostrar_snack(f"✅ Furgón '{nombre}' creado", "green700")
                self.in_furgo_nombre_nuevo.value = ""; self.in_furgo_desc_nuevo.value = ""; self.sel_furgo_tecnico.value = ""
                self.cargar_datos()
            except Exception as ex:
                self.mostrar_snack(f"❌ Error: {ex}", "red")
        hilo(_run)

    def hacer_reasignar(self, fid, dropdown):
        def _reasignar(_):
            nuevo_tec = dropdown.value or None
            def _run():
                try:
                    get_sb().table("furgones").update({"tecnico_usuario": nuevo_tec}).eq("id", fid).execute()
                    cache_invalidate("furgones_totales")
                    registrar_auditoria(estado.usuario_actual["nombre"], "REASIGNAR FURGÓN", f"ID {fid} -> {nuevo_tec}")
                    self.mostrar_snack("✅ Técnico reasignado", "green700")
                    self.cargar_datos()
                except Exception as ex:
                    self.mostrar_snack(f"❌ {ex}", "red")
            hilo(_run)
        return _reasignar

    def confirmar_eliminacion_furgo(self, fid, nombre_furgo):
        def _abrir_alerta(e):
            from utils.ui_helpers import mostrar_alerta
            mostrar_alerta(
                self.page_ref,
                "Eliminar Furgón",
                f"¿Estás seguro de que deseas eliminar el furgón '{nombre_furgo}'?\n\n"
                "Esta acción borrará el furgón y desconectará todo su stock actual. Esto no se puede deshacer.",
                lambda: self.ejecutar_eliminacion_furgo(fid, nombre_furgo),
                texto_confirmar="Sí, eliminar",
                tipo="error",
            )
        return _abrir_alerta

    def ejecutar_eliminacion_furgo(self, fid, nombre_furgo):
        def _run():
            try:
                get_sb().table("furgon_productos").delete().eq("furgon_id", fid).execute()
                get_sb().table("furgones").delete().eq("id", fid).execute()
                
                cache_invalidate("furgones_totales")
                registrar_auditoria(estado.usuario_actual["nombre"], "ELIMINAR FURGÓN", f"Furgón: {nombre_furgo}")
                self.mostrar_snack(f"🗑️ Furgón '{nombre_furgo}' eliminado correctamente", "red700")
                self.cargar_datos()
            except Exception as ex:
                self.mostrar_snack(f"❌ Error al eliminar furgón: {ex}", "red")
        hilo(_run)

    def descargar_plantilla_furgo(self, furgon_nombre):
        def _abrir(_=None):
            def _run():
                import openpyxl, io, base64 as _b64
                from openpyxl.styles import Font, PatternFill, Side, Border; from openpyxl.utils import get_column_letter
                try:
                    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Carga Stock"
                    headers = ["SKU", "Cantidad"]; fill = PatternFill("solid", fgColor="00796B"); font = Font(bold=True, color="FFFFFF"); b = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = fill; cell.font = font; cell.border = b; ws.column_dimensions[get_column_letter(col)].width = 20
                    for row_num in range(2, 502):
                        ws.cell(row=row_num, column=1).number_format = '@'
                    _buf = io.BytesIO(); wb.save(_buf)
                    import subprocess as _sp, datetime as _dt
                    _fname = f"/tmp/PlantillaCargaStock_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    open(_fname, 'wb').write(_buf.getvalue())
                    _sp.Popen(['open', _fname])
                    self.mostrar_snack(f"✅ Plantilla generada", "green700")
                except Exception as ex:
                    self.mostrar_snack(f"❌ Error al generar: {ex}", "red700")
            hilo(_run)
        return _abrir

    def hacer_buscar_adj(self, buscador_f, txt_f, cont_f, fid_f):
        def _on_seleccionar(prod):
            sku = prod["sku"]
            def _run():
                try:
                    res = get_sb().table("furgon_productos").select("cantidad, productos(nombre)").eq("furgon_id", fid_f).eq("sku", sku).execute()
                    if res.data:
                        d = res.data[0]
                        nom = d.get("productos", {}).get("nombre", sku) if isinstance(d.get("productos"), dict) else sku
                        txt_f.value = f"📦 {nom}  |  Stock furgón: {d.get('cantidad') or 0}"
                        txt_f.color = "blue800"; cont_f.bgcolor = "blue50"; cont_f.border = ft.Border(left=ft.BorderSide(3,"blue400"))
                    else:
                        txt_f.value = "⚠️ Producto no está en este furgón"
                        txt_f.color = "red700"; cont_f.bgcolor = "red50"; cont_f.border = ft.Border(left=ft.BorderSide(3,"red400"))
                    cont_f.visible = True; cont_f.update(); txt_f.update()
                except Exception as ex: self.mostrar_snack(f"❌ Error al buscar en furgón: {ex}", "red700")
            hilo(_run)
        return _on_seleccionar

    def hacer_ajuste_furgo(self, sku_f, cant_f, bod_f, txt_f, cont_f, fid_f, fname_f, delta):
        def _ajustar(_):
            sku = sku_f.value.strip()
            try: cant = int(cant_f.value or "1")
            except: cant = 1
            if not sku or cant <= 0: return
            bodega_id = bod_f.value if bod_f.value else None
            if delta > 0 and not bodega_id:
                txt_f.value = "⚠️ Selecciona una bodega de origen para sumar stock"
                txt_f.color = "orange700"; cont_f.bgcolor = "orange50"; cont_f.border = ft.Border(left=ft.BorderSide(3,"orange400"))
                cont_f.visible = True; cont_f.update(); txt_f.update(); return
            def _run():
                try:
                    res = get_sb().table("furgon_productos").select("cantidad, productos(nombre)").eq("furgon_id", fid_f).eq("sku", sku).execute()

                    def descontar_bodega(sku, bodega_id, cantidad):
                        try:
                            r = get_sb().table("bodega_productos").select("cantidad").eq("bodega_id", bodega_id).eq("sku", sku).execute()
                            if r.data:
                                nuevo = max(0, (r.data[0]["cantidad"] or 0) - cantidad)
                                get_sb().table("bodega_productos").update({"cantidad": nuevo}).eq("bodega_id", bodega_id).eq("sku", sku).execute()
                            r_total = get_sb().table("bodega_productos").select("cantidad").eq("sku", sku).execute()
                            total = sum(row.get("cantidad", 0) for row in (r_total.data or []))
                            get_sb().table("productos").update({"stock_global": total}).eq("sku", sku).execute()
                        except Exception as ex:
                            raise

                    if not res.data:
                        res_prod = get_sb().table("productos").select("nombre, stock_global").eq("sku", sku).execute()
                        if not res_prod.data:
                            txt_f.value = f"❌ SKU '{sku}' no existe en el sistema"; txt_f.color = "red700"; cont_f.bgcolor = "red50"; cont_f.border = ft.Border(left=ft.BorderSide(3,"red400"))
                            cont_f.visible = True; cont_f.update(); txt_f.update(); return
                        cantidad_inicial = max(0, delta * cant)
                        get_sb().table("furgon_productos").insert({"furgon_id": fid_f, "sku": sku, "cantidad": cantidad_inicial}).execute()
                        if delta > 0: descontar_bodega(sku, int(bodega_id), cantidad_inicial)
                        nom = res_prod.data[0].get("nombre", sku)
                        registrar_auditoria(estado.usuario_actual["nombre"], "AJUSTE FURGÓN", f"{fname_f} | SKU: {sku} | +{cantidad_inicial}")
                        txt_f.value = f"✅ {nom}  |  Agregado: {cantidad_inicial}"; txt_f.color = "green800"; cont_f.bgcolor = "green50"; cont_f.border = ft.Border(left=ft.BorderSide(3,"green500"))
                    else:
                        d = res.data[0]
                        nom = d.get("productos", {}).get("nombre", sku) if isinstance(d.get("productos"), dict) else sku
                        nuevo = max(0, (d.get("cantidad") or 0) + (delta * cant))
                        get_sb().table("furgon_productos").update({"cantidad": nuevo}).eq("furgon_id", fid_f).eq("sku", sku).execute()
                        if delta > 0: descontar_bodega(sku, int(bodega_id), cant)
                        registrar_auditoria(estado.usuario_actual["nombre"], "AJUSTE FURGÓN", f"{fname_f} | SKU: {sku} | Nuevo stock: {nuevo}")
                        txt_f.value = f"✅ {nom}  |  Stock furgón: {nuevo}"; txt_f.color = "green800"; cont_f.bgcolor = "green50"; cont_f.border = ft.Border(left=ft.BorderSide(3,"green500"))

                    cont_f.visible = True; sku_f.limpiar(); cant_f.value = "1"
                    cant_f.update(); cont_f.update(); txt_f.update()
                    self.cargar_datos()
                except Exception as ex:
                    self.mostrar_snack(f"❌ Error en ajuste: {ex}", "red700")
            hilo(_run)
        return _ajustar

    def importar_stock_furgo(self, furgon_id, furgon_nombre):
        def _abrir(_=None):
            def _on_result(e):
                if not e.files or not e.files[0].path:
                    return
                ruta = e.files[0].path
                def _run():
                    try:
                        import os as _os
                        ext = _os.path.splitext(ruta)[1].lower(); filas = []
                        if ext == ".csv":
                            import csv
                            with open(ruta, newline="", encoding="utf-8-sig") as fc:
                                for row in csv.DictReader(fc): filas.append(dict(row))
                        else:
                            import openpyxl
                            wb = openpyxl.load_workbook(ruta, data_only=True); ws = wb.active
                            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if any(v is not None for v in row): filas.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})

                        def buscar_col(row, opts):
                            for op in opts:
                                for k in row:
                                    if k and k.lower().strip() == op.lower(): return row[k]
                            return ""

                        ok = 0; err = 0
                        for fila in filas:
                            sku = buscar_col(fila, ["sku","código","codigo","cod"]); cant = buscar_col(fila, ["cantidad","stock","qty","cant"])
                            if not sku: continue
                            try:
                                if isinstance(sku, float):
                                    sku_limpio = str(int(sku)) if sku == int(sku) else str(sku)
                                elif isinstance(sku, int):
                                    sku_limpio = str(sku)
                                else:
                                    sku_limpio = str(sku).strip()
                                    if sku_limpio.endswith('.0'): sku_limpio = sku_limpio[:-2]
                                cantidad = int(float(str(cant).replace(",",".").strip())) if cant and str(cant).strip() else 0
                            except ValueError:
                                err += 1; continue
                            try:
                                res_prev = get_sb().table("furgon_productos").select("cantidad").eq("furgon_id", furgon_id).eq("sku", sku_limpio).execute()
                                if res_prev.data:
                                    cant_prev = res_prev.data[0]["cantidad"] or 0
                                    get_sb().table("furgon_productos").update({"cantidad": cant_prev + cantidad}).eq("furgon_id", furgon_id).eq("sku", sku_limpio).execute()
                                else:
                                    get_sb().table("furgon_productos").insert({"furgon_id": furgon_id, "sku": sku_limpio, "cantidad": cantidad}).execute()
                                ok += 1
                            except Exception:
                                err += 1

                        cache_invalidate("furgones_totales")
                        registrar_auditoria(estado.usuario_actual["nombre"], "IMPORTAR FURGÓN", f"{furgon_nombre}: {ok} productos importados")
                        self.txt_import_furgo.value = f"✅ {furgon_nombre}: {ok} actualizados" + (f" | ⚠️ {err} errores" if err else "")
                        self.txt_import_furgo.color = "green700" if not err else "orange700"
                        det_f = f"{ok} producto(s) importado(s)" + (f" · {err} errores" if err else "")
                        if ok > 0:
                            self.mostrar_dialogo_importacion(True, "Stock de Furgón actualizado", f"Se ha cargado el stock correctamente a '{furgon_nombre}'.", det_f)
                        else:
                            self.mostrar_dialogo_importacion(False, "Sin actualizar", "El archivo no contenía productos válidos.", det_f)
                        self.cargar_datos()
                    except Exception as ex:
                        self.mostrar_dialogo_importacion(False, "Error crítico", "Hubo un error al leer el archivo.", str(ex))
                        self.txt_import_furgo.value = f"❌ Error: {ex}"; self.txt_import_furgo.color = "red"
                        if self.page_ref: self.txt_import_furgo.update()
                hilo(_run)
            picker = ft.FilePicker(on_result=_on_result)
            self.page_ref.overlay.append(picker)
            self.page_ref.update()
            picker.pick_files(dialog_title=f"Importar stock {furgon_nombre}", allowed_extensions=["xlsx", "xls", "csv"], allow_multiple=False)
        return _abrir

    # ════════════════════════════════════════════════════════════════
    # 🛠️  HERRAMIENTAS POR FURGÓN (tabla: herramientas_furgo)
    # ════════════════════════════════════════════════════════════════
    def construir_bloque_herramientas(self, fid, fname):
        in_desc = ft.TextField(label="Descripción", expand=True, height=55)
        in_modelo = ft.TextField(label="Modelo", expand=True, height=55)
        in_stock = ft.TextField(label="Stock", value="1", width=110, height=55, text_align="center")
        in_valor = ft.TextField(label="Valor ($)", width=150, height=55)
        in_fecha = ft.TextField(label="Fecha entrega (AAAA-MM-DD)", width=220, height=55)
        txt_count = ft.Text("0 herramientas", size=12, italic=True, color="grey600")

        sel_herramienta = ft.Dropdown(label="Seleccionar herramienta", expand=True, height=55, options=[])
        in_cant_aj = ft.TextField(label="Cant.", value="1", width=90, height=55, text_align="center")
        txt_aj_info = ft.Text("", size=12, color="grey600", italic=True)
        cont_aj_info = ft.Container(content=txt_aj_info, visible=False, padding=ft.Padding(10, 6, 10, 6), border_radius=6, bgcolor="blue50", border=ft.Border(left=ft.BorderSide(3, "blue400")))

        # Guardamos referencias para poder refrescar este bloque desde otros métodos
        refs = {
            "fid": fid, "fname": fname, "sel": sel_herramienta, "txt_count": txt_count,
            "in_desc": in_desc, "in_modelo": in_modelo, "in_stock": in_stock,
            "in_valor": in_valor, "in_fecha": in_fecha,
        }
        if not hasattr(self, "_bloques_herramientas"):
            self._bloques_herramientas = {}
        self._bloques_herramientas[fid] = refs

        self.refrescar_herramientas(fid)

        bloque_agregar = ft.Container(
            content=ft.Column([
                ft.Row([ft.Icon(ft.Icons.ADD_CIRCLE, color="purple700", size=20), ft.Text("Agregar Herramienta", weight="bold", size=15, color="blue900")]),
                ft.Divider(height=10, color="transparent"),
                ft.Row([in_desc, in_modelo, in_stock, in_valor], spacing=12),
                ft.Row([in_fecha, ft.ElevatedButton("AGREGAR", icon=ft.Icons.ADD, on_click=self.hacer_agregar_herramienta(fid, fname), bgcolor="blue700", color="white", height=50)], spacing=12),
                txt_count,
            ]),
            padding=20, bgcolor=ft.Colors.with_opacity(0.08, "blue"), border_radius=12,
        )

        bloque_importar = ft.Container(
            content=ft.Column([
                ft.Row([ft.Icon(ft.Icons.FOLDER, color="amber800", size=20), ft.Text("Importar desde Excel", weight="bold", size=15, color="blue900")]),
                ft.Divider(height=8, color="transparent"),
                ft.Text("Columnas: furgon_nombre (opcional), descripcion *, modelo, stock *, valor_unitario, fecha_entrega (AAAA-MM-DD).", size=12, color="grey600"),
                ft.Divider(height=8, color="transparent"),
                ft.Row([
                    ft.OutlinedButton("↓ Descargar Plantilla", icon=ft.Icons.DOWNLOAD, on_click=self.descargar_plantilla_herramientas(fname), height=44),
                    ft.ElevatedButton("📁 Importar xlsx", icon=ft.Icons.UPLOAD_FILE, on_click=self.hacer_importar_herramientas(fid, fname), bgcolor="blue700", color="white", height=44),
                ], spacing=12),
            ]),
            padding=20, bgcolor="grey50", border_radius=12,
        )

        bloque_ajuste = ft.Container(
            content=ft.Column([
                ft.Row([ft.Text("⚖️", size=16), ft.Text("Ajuste de stock", weight="bold", size=14, color="grey800")]),
                ft.Divider(height=10, color="transparent"),
                ft.Row([sel_herramienta, in_cant_aj,
                    ft.ElevatedButton("SUMAR", icon=ft.Icons.ADD, on_click=self.hacer_ajustar_herramienta(fid, sel_herramienta, in_cant_aj, txt_aj_info, cont_aj_info, 1), bgcolor="green700", color="white", height=50),
                    ft.ElevatedButton("RESTAR", icon=ft.Icons.REMOVE, on_click=self.hacer_ajustar_herramienta(fid, sel_herramienta, in_cant_aj, txt_aj_info, cont_aj_info, -1), bgcolor="orange700", color="white", height=50),
                ], spacing=12),
                cont_aj_info,
            ]),
            padding=20, bgcolor="grey50", border_radius=12,
        )

        return ft.Container(
            content=ft.Column([
                ft.Text(f"🛠️ Herramientas — {fname}", weight="bold", size=15, color="purple900"),
                ft.Divider(height=8, color="transparent"),
                bloque_agregar,
                ft.Divider(height=12, color="transparent"),
                bloque_importar,
                ft.Divider(height=12, color="transparent"),
                bloque_ajuste,
            ]),
            padding=20, bgcolor="white", border_radius=12, border=ft.border.all(1, "grey200"),
        )

    def refrescar_herramientas(self, fid):
        refs = getattr(self, "_bloques_herramientas", {}).get(fid)
        if not refs:
            return
        def _run():
            try:
                herramientas = fetch_herramientas_furgo(fid)
                refs["sel"].options = [
                    ft.dropdown.Option(str(h["id"]), f"{h.get('descripcion','')} ({h.get('modelo') or '—'}) · Stock: {h.get('stock', 0)}")
                    for h in herramientas
                ]
                refs["txt_count"].value = f"{len(herramientas)} herramienta{'s' if len(herramientas) != 1 else ''}"
                if self.page_ref:
                    try:
                        refs["sel"].update()
                        refs["txt_count"].update()
                    except Exception:
                        pass
            except Exception as ex:
                self.mostrar_snack(f"❌ Error al refrescar herramientas: {ex}", "red700")
        hilo(_run)

    def hacer_agregar_herramienta(self, fid, fname):
        def _agregar(e):
            refs = self._bloques_herramientas[fid]
            desc = refs["in_desc"].value.strip()
            modelo = refs["in_modelo"].value.strip()
            fecha = refs["in_fecha"].value.strip()

            if not desc:
                self.mostrar_snack("⚠️ Ingresa una descripción", "orange")
                return
            try:
                stock = int(refs["in_stock"].value or "0")
            except ValueError:
                self.mostrar_snack("⚠️ Stock inválido", "orange")
                return
            try:
                valor = float((refs["in_valor"].value or "0").replace(",", "."))
            except ValueError:
                self.mostrar_snack("⚠️ Valor inválido", "orange")
                return

            fecha_valida = None
            if fecha:
                try:
                    datetime.date.fromisoformat(fecha)
                    fecha_valida = fecha
                except ValueError:
                    self.mostrar_snack("⚠️ Fecha inválida, usa AAAA-MM-DD", "orange")
                    return

            def _run():
                try:
                    get_sb().table("herramientas_furgo").insert({
                        "furgon_id": fid,
                        "descripcion": desc,
                        "modelo": modelo or None,
                        "stock": stock,
                        "valor_unitario": valor,
                        "fecha_entrega": fecha_valida,
                    }).execute()
                    registrar_auditoria(estado.usuario_actual["nombre"], "AGREGAR HERRAMIENTA", f"{fname} | {desc} | Stock: {stock}")
                    self.mostrar_snack(f"✅ Herramienta '{desc}' agregada", "green700")

                    refs["in_desc"].value = ""; refs["in_modelo"].value = ""
                    refs["in_stock"].value = "1"; refs["in_valor"].value = ""; refs["in_fecha"].value = ""
                    if self.page_ref:
                        refs["in_desc"].update(); refs["in_modelo"].update()
                        refs["in_stock"].update(); refs["in_valor"].update(); refs["in_fecha"].update()

                    self.refrescar_herramientas(fid)
                except Exception as ex:
                    self.mostrar_snack(f"❌ Error: {ex}", "red")
            hilo(_run)
        return _agregar

    def hacer_ajustar_herramienta(self, fid, sel_herramienta, in_cant, txt_info, cont_info, delta):
        def _ajustar(e):
            hid = sel_herramienta.value
            if not hid:
                self.mostrar_snack("⚠️ Selecciona una herramienta", "orange")
                return
            try:
                cant = int(in_cant.value or "1")
            except ValueError:
                cant = 1

            def _run():
                try:
                    res = get_sb().table("herramientas_furgo").select("*").eq("id", hid).execute()
                    if not res.data:
                        self.mostrar_snack("❌ Herramienta no encontrada", "red")
                        return
                    h = res.data[0]
                    nuevo = max(0, (h.get("stock") or 0) + (delta * cant))
                    get_sb().table("herramientas_furgo").update({"stock": nuevo}).eq("id", hid).execute()
                    registrar_auditoria(estado.usuario_actual["nombre"], "AJUSTE HERRAMIENTA", f"{h.get('descripcion','')} | Nuevo stock: {nuevo}")

                    txt_info.value = f"✅ {h.get('descripcion','')} | Stock: {nuevo}"
                    txt_info.color = "green800"
                    cont_info.bgcolor = "green50"
                    cont_info.border = ft.Border(left=ft.BorderSide(3, "green500"))
                    cont_info.visible = True
                    in_cant.value = "1"
                    if self.page_ref:
                        txt_info.update(); cont_info.update(); in_cant.update()

                    self.refrescar_herramientas(fid)
                except Exception as ex:
                    self.mostrar_snack(f"❌ Error: {ex}", "red")
            hilo(_run)
        return _ajustar

    def descargar_plantilla_herramientas(self, furgon_nombre):
        def _abrir(_=None):
            def _run():
                import openpyxl, io, base64 as _b64
                from openpyxl.styles import Font, PatternFill, Side, Border
                from openpyxl.utils import get_column_letter
                try:
                    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Herramientas"
                    headers = ["furgon_nombre", "descripcion", "modelo", "stock", "valor_unitario", "fecha_entrega"]
                    hints_h = ["Nombre del furgón", "Descripción de la herramienta", "Modelo o código", "Cantidad entera", "Valor en pesos CLP (entero, sin $). Ej: 15000", "Fecha AAAA-MM-DD"]
                    fill = PatternFill("solid", fgColor="6A1B9A"); font = Font(bold=True, color="FFFFFF")
                    b = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
                    from openpyxl.comments import Comment
                    for col, (h, hint) in enumerate(zip(headers, hints_h), 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = fill; cell.font = font; cell.border = b
                        ws.column_dimensions[get_column_letter(col)].width = 22
                        comment = Comment(hint, "SICAV ERP"); comment.width = 220; comment.height = 50
                        cell.comment = comment
                    for row_num in range(2, 502):
                        ws.cell(row=row_num, column=5).number_format = "#,##0"
                    _buf = io.BytesIO(); wb.save(_buf)
                    import subprocess as _sp, datetime as _dt
                    _fname = f"/tmp/PlantillaHerramientas_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    open(_fname, 'wb').write(_buf.getvalue())
                    _sp.Popen(['open', _fname])
                    self.mostrar_snack("✅ Plantilla generada", "green700")
                except Exception as ex:
                    self.mostrar_snack(f"❌ Error al generar: {ex}", "red700")
            hilo(_run)
        return _abrir

    def hacer_importar_herramientas(self, fid, fname):
        def _abrir(_=None):
            def _on_result(e):
                if not e.files or not e.files[0].path:
                    return
                ruta = e.files[0].path
                def _run():
                    try:
                        import os as _os
                        ext = _os.path.splitext(ruta)[1].lower(); filas = []
                        if ext == ".csv":
                            import csv
                            with open(ruta, newline="", encoding="utf-8-sig") as fc:
                                for row in csv.DictReader(fc): filas.append(dict(row))
                        else:
                            import openpyxl
                            wb = openpyxl.load_workbook(ruta, data_only=True); ws = wb.active
                            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if any(v is not None for v in row):
                                    filas.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})

                        def buscar_col(row, opts):
                            for op in opts:
                                for k in row:
                                    if k and k.strip().lower() == op.lower(): return row[k]
                            return ""

                        res_existentes = get_sb().table("herramientas_furgo").select("id, descripcion, modelo, stock, valor_unitario, fecha_entrega").eq("furgon_id", fid).execute()
                        dict_existentes = {}
                        for h in (res_existentes.data or []):
                            desc_key = str(h["descripcion"]).strip().lower()
                            mod_key = str(h["modelo"]).strip().lower() if h.get("modelo") else ""
                            dict_existentes[(desc_key, mod_key)] = h

                        ok = 0; err = 0
                        for fila in filas:
                            desc = buscar_col(fila, ["descripcion", "descripción"])
                            if not desc: err += 1; continue
                            modelo = buscar_col(fila, ["modelo"])
                            stock_raw = buscar_col(fila, ["stock", "cantidad"])
                            valor_raw = buscar_col(fila, ["valor_unitario", "valor"])
                            fecha_raw = buscar_col(fila, ["fecha_entrega", "fecha"])
                            try:
                                stock = int(float(str(stock_raw).replace(",", "."))) if stock_raw and str(stock_raw).strip() else 0
                                valor = float(str(valor_raw).replace(",", ".")) if valor_raw and str(valor_raw).strip() else 0
                                fecha_valida = None
                                if fecha_raw:
                                    try: fecha_valida = datetime.date.fromisoformat(str(fecha_raw).strip()).isoformat()
                                    except: fecha_valida = None
                            except: err += 1; continue
                            desc_limpio = str(desc).strip()
                            modelo_limpio = str(modelo).strip() if modelo else ""
                            key = (desc_limpio.lower(), modelo_limpio.lower())
                            if key in dict_existentes:
                                h_existente = dict_existentes[key]
                                nuevo_stock = (h_existente.get("stock") or 0) + stock
                                get_sb().table("herramientas_furgo").update({
                                    "stock": nuevo_stock,
                                    "valor_unitario": valor if valor > 0 else h_existente.get("valor_unitario", 0),
                                    "fecha_entrega": fecha_valida if fecha_valida else h_existente.get("fecha_entrega")
                                }).eq("id", h_existente["id"]).execute()
                                dict_existentes[key]["stock"] = nuevo_stock
                            else:
                                res_ins = get_sb().table("herramientas_furgo").insert({
                                    "furgon_id": fid,
                                    "descripcion": desc_limpio,
                                    "modelo": modelo_limpio if modelo_limpio else None,
                                    "stock": stock,
                                    "valor_unitario": valor,
                                    "fecha_entrega": fecha_valida,
                                }).execute()
                                if res_ins.data:
                                    dict_existentes[key] = res_ins.data[0]
                            ok += 1

                        registrar_auditoria(estado.usuario_actual["nombre"], "IMPORTAR HERRAMIENTAS", f"{fname}: {ok} herramientas procesadas")
                        self.refrescar_herramientas(fid)
                        if self.page_ref:
                            self.mostrar_dialogo_importacion(ok > 0, "Proceso finalizado", f"Se procesaron {ok} herramientas correctamente.", f"{err} errores detectados.")
                    except Exception as ex:
                        self.mostrar_dialogo_importacion(False, "Error crítico", "No se pudo mapear el archivo.", str(ex))
                hilo(_run)
            picker = ft.FilePicker(on_result=_on_result)
            self.page_ref.overlay.append(picker)
            self.page_ref.update()
            picker.pick_files(dialog_title=f"Importar herramientas {fname}", allowed_extensions=["xlsx", "xls", "csv"], allow_multiple=False)
        return _abrir

    def exportar_pdf_furgones(self, e):
        self.mostrar_snack("⏳ Generando reporte PDF, por favor espera...", "blue700")

        def _run_pdf_generation():
            try:
                from fpdf import FPDF
                
                def txt(t): return t.encode('latin-1', 'replace').decode('latin-1')
                
                pdf = FPDF(orientation='P', unit='mm', format='A4')
                pdf.add_page()
                pdf.set_font("Arial", "B", 16)
                pdf.cell(190, 10, txt("Reporte Global de Furgones"), ln=1, align="C")
                pdf.ln(5)
                
                res_f = get_sb().table("furgones").select("*").execute()
                furgones = res_f.data or []
                
                for f in furgones:
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(190, 10, txt(f"Furgón: {f['nombre']} | Técnico: {f.get('tecnico_usuario', 'Sin asignar')}"), ln=1)
                    
                    pdf.set_font("Arial", "B", 10)
                    pdf.cell(160, 8, txt("Herramienta"), border=1)
                    pdf.cell(30, 8, "Stock", border=1, align="C")
                    pdf.ln(8)

                    res_h = get_sb().table("herramientas_furgo").select("descripcion, stock").eq("furgon_id", f["id"]).execute()
                    herramientas = res_h.data or []
                    
                    pdf.set_font("Arial", "", 9)
                    if not herramientas:
                        pdf.cell(190, 8, txt("No hay herramientas asignadas a este furgón."), border=1, align="C", ln=1)
                    else:
                        for h in herramientas:
                            desc = txt(h.get("descripcion", "")[:90])
                            cant = str(h.get("stock", 0))
                            
                            pdf.cell(160, 8, desc, border=1)
                            pdf.cell(30, 8, cant, border=1, align="C")
                            pdf.ln(8)
                    pdf.ln(5)

                import subprocess as _sp, datetime as _dt
                _fname = f"/tmp/ReporteHerramientas_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                open(_fname, 'wb').write(pdf.output())
                _sp.Popen(['open', _fname])
                self.mostrar_snack(f"✅ Reporte generado con éxito.", "green700")
            except Exception as ex:
                self.mostrar_snack(f"❌ Error al generar PDF: {ex}", "red700")

        hilo(_run_pdf_generation)

    def exportar_pdf_inventario_furgon(self, e):
        self.mostrar_snack("⏳ Generando reporte PDF, por favor espera...", "blue700")
        def _run():
            try:
                from fpdf import FPDF
            except ImportError:
                self.mostrar_snack("❌ Falta la librería fpdf. Ejecuta en la terminal: pip install fpdf", "red700")
                return
            try:
                # 2. Cargar diccionario de productos
                res_prod = get_sb().table("productos").select("sku, nombre").execute()
                mapa_productos = {p["sku"]: p["nombre"] for p in (res_prod.data or [])}

                # 3. Inicializar el documento PDF
                pdf = FPDF()
                pdf.set_auto_page_break(auto=True, margin=15)
                
                # 4. Traer todos los furgones
                res_f = get_sb().table("furgones").select("*").order("nombre").execute()
                furgones = res_f.data or []

                # --- FILTRO ANTI-ERRORES PARA FPDF ---
                def txt(texto):
                    if not texto: return ""
                    return str(texto).encode('latin-1', 'replace').decode('latin-1')

                if not furgones:
                    pdf.add_page()
                    pdf.set_font("Arial", "B", 14)
                    pdf.cell(0, 10, txt("No hay furgones registrados en el sistema."), ln=1, align="C")
                else:
                    for f in furgones:
                        pdf.add_page()
                        
                        # --- ENCABEZADO DEL FURGÓN ---
                        pdf.set_font("Arial", "B", 16)
                        pdf.set_fill_color(21, 101, 192) # Color Azul Supabase
                        pdf.set_text_color(255, 255, 255)
                        pdf.cell(0, 12, txt(f"INVENTARIO: {f['nombre'].upper()}"), ln=1, fill=True, align="C")
                        
                        pdf.ln(5)
                        pdf.set_font("Arial", "B", 12)
                        pdf.set_text_color(50, 50, 50)
                        pdf.cell(0, 8, txt(f"Técnico Asignado: {f.get('tecnico_usuario') or 'Sin asignar'}"), ln=1)
                        pdf.set_font("Arial", "", 11)
                        pdf.cell(0, 6, txt(f"Patente: {f.get('patente') or 'N/A'}  |  Descripción: {f.get('descripcion') or 'N/A'}"), ln=1)
                        pdf.ln(10)

                        # --- SECCIÓN 1: REPUESTOS (STOCK) ---
                        pdf.set_font("Arial", "B", 12)
                        pdf.set_text_color(0, 0, 0)
                        pdf.cell(0, 8, txt("1. LISTADO DE REPUESTOS"), ln=1)
                        
                        # Cabecera
                        pdf.set_font("Arial", "B", 10)
                        pdf.set_fill_color(240, 240, 240)
                        pdf.cell(40, 8, "SKU", border=1, fill=True)
                        pdf.cell(120, 8, "Nombre del Producto", border=1, fill=True)
                        pdf.cell(30, 8, "Cantidad", border=1, fill=True, align="C")
                        pdf.ln(8)

                        res_p = get_sb().table("furgon_productos").select("sku, cantidad, productos(nombre)").eq("furgon_id", f["id"]).execute()
                        repuestos = res_p.data or []

                        pdf.set_font("Arial", "", 9)
                        if not repuestos:
                            pdf.cell(190, 8, txt("No hay repuestos registrados en este furgón."), border=1, align="C", ln=1)
                        else:
                            for r in repuestos:
                                sku = txt(r["sku"])
                                cant = str(r["cantidad"])
                                prod_obj = r.get("productos")
                                nombre_raw = (prod_obj.get("nombre") if isinstance(prod_obj, dict) else None) or mapa_productos.get(r["sku"], r["sku"])
                                nombre = txt(nombre_raw[:65])
                                
                                pdf.cell(40, 8, sku, border=1)
                                pdf.cell(120, 8, nombre, border=1)
                                pdf.cell(30, 8, cant, border=1, align="C")
                                pdf.ln(8)
                        
                        pdf.ln(10)

                        # --- SECCIÓN 2: HERRAMIENTAS ---
                        pdf.set_font("Arial", "B", 12)
                        pdf.cell(0, 8, txt("2. INVENTARIO DE HERRAMIENTAS"), ln=1)
                        
                        pdf.set_font("Arial", "B", 10)
                        pdf.set_fill_color(240, 240, 240)
                        pdf.cell(160, 8, txt("Descripción / N° Serie"), border=1, fill=True)
                        pdf.cell(30, 8, "Cantidad", border=1, fill=True, align="C")
                        pdf.ln(8)

                        # CORRECCIÓN: Se cambió "cantidad" por "stock" en la consulta a la BD
                        res_h = get_sb().table("herramientas_furgo").select("descripcion, stock").eq("furgon_id", f["id"]).execute()
                        herramientas = res_h.data or []
                        
                        pdf.set_font("Arial", "", 9)
                        if not herramientas:
                            pdf.cell(190, 8, txt("No hay herramientas asignadas a este furgón."), border=1, align="C", ln=1)
                        else:
                            for h in herramientas:
                                desc = txt(h.get("descripcion", "")[:90])
                                # CORRECCIÓN: Se lee la llave "stock"
                                cant = str(h.get("stock", 0))
                                
                                pdf.cell(160, 8, desc, border=1)
                                pdf.cell(30, 8, cant, border=1, align="C")
                                pdf.ln(8)

                import subprocess as _sp, datetime as _dt
                _fname = f"/tmp/InventarioFurgon_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                open(_fname, 'wb').write(pdf.output())
                _sp.Popen(['open', _fname])
                self.mostrar_snack(f"✅ Reporte generado con éxito.", "green700")

            except Exception as ex:
                self.mostrar_snack(f"❌ Error al generar PDF: {ex}", "red700")

        hilo(_run)