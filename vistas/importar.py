# vistas/importar.py
import flet as ft
import os
from core.database import get_sb
from core.estado import estado
from core.utilidades import hilo, registrar_auditoria


def _label_bodega(b: dict) -> str:
    area = (b.get("areas") or {}).get("nombre") or ""
    return f"[{area}] {b['nombre']}" if area else b["nombre"]

class VistaImportar(ft.Container):
    def __init__(self, page: ft.Page):
        super().__init__()
        self.page_ref = page
        self.expand = True
        self.archivo_path = None
        self._picker_abrir = ft.FilePicker(on_result=self._on_archivo_result)
        page.overlay.append(self._picker_abrir)
        self.construir_ui()

    def mostrar_snack(self, mensaje, tipo_o_color="success"):
        from utils.ui_helpers import mostrar_snack as _snack
        _snack(self.page_ref, mensaje, tipo_o_color)

    def construir_ui(self):
        # ── Selector de Bodegas ──
        self.dd_bodegas = ft.Dropdown(
            label="Seleccionar Bodega de Destino",
            width=300
        )

        self.btn_cargar_bodegas = ft.ElevatedButton(
            "Cargar Bodegas",
            icon=ft.Icons.REFRESH,
            tooltip="Recargar lista de bodegas disponibles",
            on_click=lambda _: self.cargar_lista_bodegas()
        )

        self.txt_ruta = ft.Text("Ningún archivo seleccionado", color="grey500", italic=True, expand=True)
        self.btn_procesar = ft.ElevatedButton(
            "🚀 Procesar e Importar",
            icon=ft.Icons.UPLOAD,
            tooltip="Iniciar importación masiva del archivo Excel seleccionado",
            on_click=self.procesar_excel,
            disabled=True,
            bgcolor="blue800",
            color="white",
            height=46,
        )
        self.progress = ft.ProgressBar(visible=False, color="blue800")
        self.col_logs = ft.Column([], scroll=ft.ScrollMode.AUTO, height=280, spacing=2)

        # ── Header ──
        header = ft.Column([
            ft.Text("📥 Importación Masiva de Inventario", size=24, weight="bold", color="blue900"),
            ft.Text(
                "Sube un Excel (.xlsx) con columnas: sku*, nombre*, familia, subfamilia, costo_neto, precio_venta, stock",
                size=13, color="grey600",
            ),
        ])

        # ── Body ──
        body = ft.Column([
            ft.Divider(height=10, color="transparent"),

            # ── Panel de selección ──
            ft.Container(
                content=ft.Column([
                    ft.Row([
                        self.dd_bodegas,
                        self.btn_cargar_bodegas
                    ], spacing=12),
                    ft.Divider(height=8, color="transparent"),
                    ft.Row([
                        ft.ElevatedButton(
                            "📂 Seleccionar Excel",
                            icon=ft.Icons.FOLDER_OPEN,
                            tooltip="Abrir explorador para seleccionar archivo Excel",
                            on_click=self.abrir_selector,
                            height=46,
                        ),
                        self.txt_ruta,
                    ], spacing=12),
                    ft.Divider(height=8, color="transparent"),
                    ft.Row([
                        self.btn_procesar,
                        ft.OutlinedButton(
                            "📄 Descargar Plantilla",
                            icon=ft.Icons.DOWNLOAD,
                            tooltip="Descargar plantilla Excel con el formato correcto",
                            on_click=self.descargar_plantilla,
                            height=46,
                        ),
                    ], spacing=12),
                    self.progress,
                ]),
                padding=20, bgcolor="white", border_radius=12, border=ft.border.all(1, "grey200"),
            ),

            ft.Divider(height=10, color="transparent"),

            # ── Log ──
            ft.Text("📋 Registro de importación:", weight="bold", size=13, color="grey700"),
            ft.Container(
                content=self.col_logs,
                padding=ft.Padding(12, 8, 12, 8),
                bgcolor="#1a1a2e",
                border_radius=10,
                expand=True,
            ),
        ], expand=True, scroll=ft.ScrollMode.AUTO)

        from utils.ui_helpers import sandwich
        self.padding = 0
        self.content = sandwich(header, body)

    def pre_cargar_si_cache_listo(self):
        from core.cache import app_cache
        if getattr(self, '_bodegas_listas', False) or not app_cache.inv_ready.is_set():
            return
        bodegas = app_cache.inv_bodegas or []
        if bodegas:
            self.dd_bodegas.options = [ft.dropdown.Option(str(b["id"]), b["nombre"]) for b in bodegas]
            self._bodegas_listas = True
            print(f"[IMPORTAR] bodegas pre-cargadas ({len(bodegas)})")

    def inicializar(self):
        if getattr(self, '_bodegas_listas', False):
            return
        self.cargar_lista_bodegas()

    def cargar_lista_bodegas(self):
        def _run():
            try:
                sb = get_sb()
                res = sb.table("bodegas").select("id, nombre, areas(nombre)").execute()
                if res.data:
                    self.dd_bodegas.options = [ft.dropdown.Option(str(b["id"]), _label_bodega(b)) for b in res.data]
                    if self.page_ref:
                        self.dd_bodegas.update()
                    self.mostrar_snack("✅ Bodegas cargadas", "success")
            except Exception as e:
                self.mostrar_snack(f"❌ Error cargando bodegas: {e}", "error")
        hilo(_run)

    def _on_archivo_result(self, e: ft.FilePickerResultEvent):
        if e.files:
            ruta = e.files[0].path
            if ruta:
                self.archivo_path = ruta
                self.txt_ruta.value = os.path.basename(ruta)
                self.txt_ruta.color = "blue700"
                self.btn_procesar.disabled = False
            else:
                self.archivo_path = None
                self.btn_procesar.disabled = True
                self.txt_ruta.value = "Ningún archivo seleccionado"
                self.txt_ruta.color = "grey500"
        else:
            self.archivo_path = None
            self.btn_procesar.disabled = True
            self.txt_ruta.value = "Ningún archivo seleccionado"
            self.txt_ruta.color = "grey500"
        if self.page_ref:
            self.txt_ruta.update()
            self.btn_procesar.update()

    def abrir_selector(self, e):
        self._picker_abrir.pick_files(
            dialog_title="Seleccionar Excel de inventario",
            allowed_extensions=["xlsx", "xls"],
            allow_multiple=False,
        )

    def descargar_plantilla(self, e):
        def _run():
            import openpyxl, io
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.comments import Comment
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Inventario"
                ws.freeze_panes = "A2"

                headers = ["sku", "nombre", "familia", "subfamilia", "costo_neto", "precio_venta", "stock"]
                hints   = [
                    "TEXTO obligatorio. Ej: 1234 o ABC-01. No usar decimales.",
                    "Nombre del producto (obligatorio)",
                    "Familia (se crea si no existe). Dejar en blanco = Sin Familia",
                    "Subfamilia (opcional, debe pertenecer a la familia)",
                    "Costo neto en pesos chilenos (CLP), entero. Ej: 1500",
                    "Precio de venta en pesos chilenos (CLP), entero. Ej: 2000",
                    "Stock inicial entero. Ej: 10",
                ]
                fill_hdr  = PatternFill("solid", fgColor="1565C0")
                fill_req  = PatternFill("solid", fgColor="E3F2FD")
                font_hdr  = Font(bold=True, color="FFFFFF", size=11)
                font_data = Font(size=11)
                border    = Border(
                    left=Side(style="thin", color="BDBDBD"), right=Side(style="thin", color="BDBDBD"),
                    top=Side(style="thin", color="BDBDBD"),  bottom=Side(style="thin", color="BDBDBD"),
                )
                align_c = Alignment(horizontal="center", vertical="center")
                align_l = Alignment(horizontal="left",   vertical="center")
                col_widths = [18, 36, 20, 20, 16, 16, 10]

                for col, (h, hint, w) in enumerate(zip(headers, hints, col_widths), 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = fill_hdr; cell.font = font_hdr; cell.border = border; cell.alignment = align_c
                    ws.column_dimensions[get_column_letter(col)].width = w
                    comment = Comment(hint, "SICAV ERP"); comment.width = 260; comment.height = 60
                    cell.comment = comment
                ws.row_dimensions[1].height = 22

                ejemplos = [
                    ["1001",    "Válvula de bola 1/2\"", "Válvulas",    "Bola",   2500, 3800, 5],
                    ["1002",    "Filtro de agua 3/4\"",  "Filtros",     "Agua",   1800, 2700, 12],
                    ["TOOL-01", "Llave inglesa 12\"",    "Herramientas","Llaves", 4500, 6900, 3],
                ]
                for r_idx, row_data in enumerate(ejemplos, 2):
                    for c_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=val)
                        cell.font = font_data; cell.border = border
                        cell.alignment = align_c if c_idx != 2 else align_l
                        if c_idx <= 2:
                            cell.fill = fill_req

                for row in ws.iter_rows(min_row=2, max_row=1000, min_col=1, max_col=1):
                    for cell in row: cell.number_format = "@"
                for row in ws.iter_rows(min_row=2, max_row=1000, min_col=5, max_col=6):
                    for cell in row: cell.number_format = "#,##0"
                for row in ws.iter_rows(min_row=2, max_row=1000, min_col=7, max_col=7):
                    for cell in row: cell.number_format = "#,##0"

                ws2 = wb.create_sheet("Instrucciones")
                instrucciones = [
                    ("INSTRUCCIONES DE USO", True),
                    ("", False),
                    ("1. La columna 'sku' es OBLIGATORIA y debe ser texto (no número decimal).", False),
                    ("   Si el SKU es numérico (ej: 1234), escribirlo sin decimales.", False),
                    ("", False),
                    ("2. La columna 'nombre' es OBLIGATORIA.", False),
                    ("", False),
                    ("3. 'familia' y 'subfamilia' son opcionales. Si la familia no existe, se creará.", False),
                    ("", False),
                    ("4. 'costo_neto' y 'precio_venta' son valores en pesos CLP (sin $). Ej: 1500", False),
                    ("", False),
                    ("5. 'stock' debe ser número entero. Ej: 10", False),
                    ("", False),
                    ("6. No modificar los nombres de las columnas (fila 1).", False),
                    ("7. Puedes agregar más filas desde la fila 2 en adelante.", False),
                    ("8. Filas completamente vacías son ignoradas.", False),
                ]
                ws2.column_dimensions["A"].width = 80
                for i, (txt, bold) in enumerate(instrucciones, 1):
                    c = ws2.cell(row=i, column=1, value=txt)
                    c.font = Font(bold=bold, size=11 if not bold else 13, color="1565C0" if bold else "000000")

                _buf = io.BytesIO()
                wb.save(_buf)
                import subprocess as _sp, datetime as _dt
                _fname = f"/tmp/PlantillaProductos_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                open(_fname, 'wb').write(_buf.getvalue())
                _sp.Popen(['open', _fname])
                self.mostrar_snack("✅ Plantilla generada", "success")
            except Exception as ex:
                self.mostrar_snack(f"❌ Error: {ex}", "error")
        hilo(_run)

    # ── Logger en el panel negro ──
    def log(self, texto, color="white"):
        self.col_logs.controls.append(
            ft.Text(texto, color=color, size=11, font_family="Consolas")
        )
        if len(self.col_logs.controls) > 500:
            self.col_logs.controls.pop(0)
        if self.page_ref:
            try: self.col_logs.update()
            except Exception: pass

    # ── Procesamiento principal ──
    def procesar_excel(self, e):
        bodega_id_str = self.dd_bodegas.value
        if not bodega_id_str:
            self.mostrar_snack("❌ Debes seleccionar una bodega de destino primero.", "error")
            return

        bodega_id = int(bodega_id_str)

        if not self.archivo_path:
            return

        self.btn_procesar.disabled = True
        self.progress.visible = True
        self.col_logs.controls.clear()
        self.log(f"⏳ Leyendo archivo para cargar en Bodega ID {bodega_id}…", "yellow")
        if self.page_ref:
            self.btn_procesar.update()
            self.progress.update()
            self.col_logs.update()

        def _run():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(self.archivo_path, data_only=True)
                ws = wb.active

                # Normalizar cabeceras
                headers = [
                    str(c.value).strip().lower().replace(" ", "_") if c.value else ""
                    for c in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                self.log(f"📋 Columnas detectadas: {headers}", "cyan")

                sb = get_sb()

                # ── Caché local para evitar N+1 ──
                cache_familias = {}
                cache_subfamilias = {}

                ok = 0; err = 0; omitidos = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in row):
                        continue
                    def _str(v):
                        if isinstance(v, float) and v.is_integer():
                            return str(int(v))
                        return str(v).strip() if v is not None else ""
                    fila = {headers[i]: _str(v) for i, v in enumerate(row) if i < len(headers)}

                    def get_col(*opts):
                        for op in opts:
                            v = fila.get(op, "").strip()
                            if v and v.lower() != "nan":
                                return v
                        return ""

                    sku = get_col("sku", "código", "codigo", "cod")
                    nombre = get_col("nombre", "descripcion", "descripción", "producto")
                    if not sku:
                        omitidos += 1; continue

                    try:
                        # ── Familia ──
                        fam_str = get_col("familia", "family") or "Sin Familia"
                        if fam_str not in cache_familias:
                            res_f = sb.table("familias").select("id").eq("nombre", fam_str).execute()
                            if res_f.data:
                                cache_familias[fam_str] = res_f.data[0]["id"]
                            else:
                                cache_familias[fam_str] = sb.table("familias").insert({"nombre": fam_str}).execute().data[0]["id"]
                                self.log(f"📁 Nueva familia: {fam_str}", "cyan")
                        fam_id = cache_familias[fam_str]

                        # ── Subfamilia ──
                        sub_str = get_col("subfamilia", "subfamilia_nombre", "subfamily")
                        sub_id = None
                        if sub_str:
                            cache_key = f"{fam_id}:{sub_str}"
                            if cache_key not in cache_subfamilias:
                                res_s = sb.table("subfamilias").select("id").eq("familia_id", fam_id).eq("nombre", sub_str).execute()
                                if res_s.data:
                                    cache_subfamilias[cache_key] = res_s.data[0]["id"]
                                else:
                                    cache_subfamilias[cache_key] = sb.table("subfamilias").insert({"familia_id": fam_id, "nombre": sub_str}).execute().data[0]["id"]
                                    self.log(f"   ↳ Nueva subfamilia: {sub_str}", "cyan")
                            sub_id = cache_subfamilias[cache_key]

                        # ── Costo / Precio / Stock ──
                        def to_num(k, default=0.0):
                            try:
                                return float(str(fila.get(k, default)).replace(",", ".").replace("$", "").strip() or default)
                            except ValueError:
                                return default

                        tiene_costo  = any(h in headers for h in ["costo_neto"])
                        tiene_precio = any(h in headers for h in ["precio_venta"])
                        tiene_stock  = any(h in headers for h in ["stock"])
                        tiene_fam    = any(h in headers for h in ["familia", "family"])
                        tiene_sub    = any(h in headers for h in ["subfamilia", "subfamily", "subfamilia_nombre"])

                        costo = to_num("costo_neto", 0) if tiene_costo else None
                        precio = to_num("precio_venta", 0) if tiene_precio else None
                        stock = int(to_num("stock", 0)) if tiene_stock else None

                        # ── Upsert Producto: solo campos presentes en el archivo ──
                        datos_prod = {"sku": sku}
                        if nombre:                       datos_prod["nombre"]       = nombre
                        if tiene_fam:                    datos_prod["familia_id"]   = fam_id
                        if tiene_sub:                    datos_prod["subfamilia_id"]= sub_id
                        if tiene_costo:                  datos_prod["costo_neto"]   = costo
                        if tiene_precio:                 datos_prod["precio_venta"] = precio
                        if tiene_stock:                  datos_prod["stock_global"] = stock

                        sb.table("productos").upsert(datos_prod, on_conflict="sku").execute()

                        # ── Upsert en Bodega: solo si el archivo trae stock ──
                        if tiene_stock:
                            sb.table("bodega_productos").upsert(
                                {"bodega_id": bodega_id, "sku": sku, "cantidad": stock},
                                on_conflict="bodega_id,sku",
                            ).execute()

                        ok += 1
                        if ok % 50 == 0:
                            self.log(f"   … {ok} productos procesados", "grey400")

                    except Exception as ex_row:
                        err += 1
                        self.log(f"❌ Error en SKU '{sku}': {ex_row}", "red")

                registrar_auditoria(
                    estado.usuario_actual.get("nombre", "Sistema"),
                    "IMPORTACION MASIVA",
                    f"{ok} OK · {err} errores · {omitidos} omitidos · Bodega: {bodega_id}",
                )
                self.log("", "transparent")
                self.log(f"🏁 Finalizado: {ok} importados · {err} con error · {omitidos} filas vacías", "yellow")
                self.mostrar_snack(f"✅ {ok} productos importados a la bodega", "success")
                from core.cache import app_cache
                app_cache.inv_ready.clear()

            except Exception as ex:
                self.log(f"❌ Error crítico: {ex}", "red")
                self.mostrar_snack(f"❌ {ex}", "error")
            finally:
                self.btn_procesar.disabled = False
                self.progress.visible = False
                if self.page_ref:
                    self.btn_procesar.update()
                    self.progress.update()
        hilo(_run)
