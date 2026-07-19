# vistas/graficos.py
import flet as ft
import os
import tempfile
import datetime
from collections import Counter, defaultdict
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

from core.database import get_sb, fetch_all, fetch_furgones_con_totales
from core.estado import estado
from core.utilidades import hilo, registrar_auditoria

_ASSETS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets")
os.makedirs(_ASSETS_DIR, exist_ok=True)

try:
    from flet.core.alignment import center as align_center
except ImportError:
    align_center = "center"

# ── Catálogo de gráficos agrupados ──────────────────────────────────────────
GRUPOS_GRAFICOS = {
    "📦 Inventario y Stock": [
        ("stock_familia",      "Stock total por familia"),
        ("stock_subfamilia",   "Stock total por subfamilia"),
        ("stock_bodega",       "Stock total por bodega"),
        ("salud_stock",        "Salud del stock (sano / crítico / cero)"),
        ("top_stock",          "Top 10 productos con más stock"),
        ("top_sin_stock",      "Familias con más productos sin stock"),
    ],
    "💰 Finanzas y Márgenes": [
        ("valor_familia",      "Valor de inventario por familia (costo vs venta)"),
        ("valor_bodega",       "Valor de inventario por bodega"),
        ("margen_familia",     "Margen promedio por familia (%)"),
        ("top_margen_prod",    "Top 10 productos por margen absoluto ($)"),
        ("anomalias",          "Anomalías financieras"),
        ("dispersion",         "Dispersión: costo vs precio de venta"),
    ],
    "🔄 Movimientos": [
        ("movimientos_tipo",   "Movimientos por tipo"),
        ("movimientos_mes",    "Evolución mensual de movimientos"),
        ("top_despachos",      "Top 10 productos más despachados a terreno"),
    ],
    "📊 Distribución": [
        ("dist_familia",       "Distribución de productos por familia"),
        ("dist_subfamilia",    "Distribución de productos por subfamilia"),
        ("comparativa_bd_fg",  "Comparativa stock: bodegas vs furgones"),
        ("furgon_stock",       "Unidades totales por furgón"),
        ("furgon_valor",       "Valor por furgón (costo vs venta)"),
    ],
    "👤 Usuarios y Actividad": [
        ("top_usuarios",       "Top 10 usuarios más activos"),
        ("acciones_tipo",      "Distribución de acciones en auditoría"),
    ],
}

# Lista plana para referencias rápidas
TIPOS_GRAFICOS = [(k, v) for grupo in GRUPOS_GRAFICOS.values() for k, v in grupo]
_TIPO_LABEL = dict(TIPOS_GRAFICOS)

_DESCRIPCIONES = {
    "stock_familia":     "Stock total acumulado por familia de productos.",
    "stock_subfamilia":  "Unidades en stock desglosadas por subfamilia.",
    "stock_bodega":      "Total de unidades almacenadas en cada bodega física.",
    "salud_stock":       "Clasifica todos los productos en sano (>3 ud.), crítico (1-3 ud.) y sin stock (0). El número central es el total de productos.",
    "top_stock":         "Los 10 productos con mayor cantidad de unidades disponibles.",
    "top_sin_stock":     "Familias con más productos en cero stock — útil para priorizar reabastecimiento.",
    "valor_familia":     "Valor económico del inventario por familia. Azul = costo neto, verde = precio de venta.",
    "valor_bodega":      "Valor del inventario en cada bodega (costo vs precio de venta).",
    "margen_familia":    "Margen de ganancia promedio (%) de los productos de cada familia.",
    "top_margen_prod":   "Top 10 productos con mayor margen absoluto por unidad (precio − costo).",
    "anomalias":         "Productos sin precio registrado, sin costo, o que se venden a pérdida.",
    "dispersion":        "Cada punto es un producto. Verde = precio > costo, rojo = vende a pérdida. La línea es el punto de equilibrio.",
    "movimientos_tipo":  "Proporción de movimientos según tipo (despacho técnico, traspaso, ajuste, venta, etc.).",
    "movimientos_mes":   "Evolución mensual del total de movimientos de inventario (últimos 18 meses).",
    "top_despachos":     "Los 10 productos más despachados a terreno por técnicos.",
    "dist_familia":      "Distribución del catálogo de productos por familia (cantidad de SKUs).",
    "dist_subfamilia":   "Distribución del catálogo por subfamilia.",
    "comparativa_bd_fg": "Comparación del stock total entre bodegas físicas y furgones.",
    "furgon_stock":      "Unidades totales cargadas actualmente en cada furgón.",
    "furgon_valor":      "Valor del inventario en cada furgón (costo vs venta).",
    "top_usuarios":      "Usuarios con más acciones registradas en el historial de auditoría.",
    "acciones_tipo":     "Distribución de los tipos de acciones en el registro de auditoría.",
}

_FIGSIZE_PIE = {"salud_stock", "movimientos_tipo", "dist_familia", "dist_subfamilia"}


class VistaGraficos(ft.Container):
    def __init__(self, page: ft.Page):
        super().__init__()
        self.page_ref = page
        self.expand = True
        self._dir_tmp = os.path.join(tempfile.gettempdir(), "sicav_graficos")
        os.makedirs(self._dir_tmp, exist_ok=True)

        self.datos_productos   = []
        self.datos_furgones    = []
        self.datos_auditoria   = []
        self.datos_movimientos = []
        self.datos_bp          = []   # bodega_productos
        self.datos_fp          = []   # furgon_productos
        self.datos_bodegas     = []   # tabla bodegas
        self.familias_unicas   = ["Todas"]

        self.construir_ui()

    def mostrar_snack(self, mensaje, tipo_o_color="success"):
        from utils.ui_helpers import mostrar_snack as _snack
        _snack(self.page_ref, mensaje, tipo_o_color)

    # ══════════════════════════════ UI ══════════════════════════════
    def construir_ui(self):
        ver_costo = estado.puede_ver("costo")

        # Filtrar grupos según permisos
        opciones = []
        primer_valido = None
        for grupo, items in GRUPOS_GRAFICOS.items():
            if "Finanzas" in grupo and not ver_costo:
                continue
            opciones.append(ft.dropdown.Option(key=f"__sep__{grupo}", text=f"── {grupo} ──"))
            for k, v in items:
                opciones.append(ft.dropdown.Option(key=k, text=f"    {v}"))
                if primer_valido is None:
                    primer_valido = k

        self.dd_tipo = ft.Dropdown(
            label="Tipo de gráfico",
            options=opciones,
            value=primer_valido or "stock_familia",
            expand=3, height=55,
        )
        self.dd_tipo.on_change = lambda _: self._actualizar_preview()
        self.dd_familia = ft.Dropdown(
            label="Filtrar familia",
            options=[ft.dropdown.Option("Todas")],
            value="Todas",
            expand=1, height=55,
        )
        self.dd_familia.on_change = lambda _: self._actualizar_preview()
        self.dd_bodega = ft.Dropdown(
            label="Filtrar bodega",
            options=[ft.dropdown.Option("Todas")],
            value="Todas",
            expand=1, height=55,
        )
        self.dd_bodega.on_change = lambda _: self._actualizar_preview()
        self.dd_furgon = ft.Dropdown(
            label="Filtrar furgón",
            options=[ft.dropdown.Option("Todos")],
            value="Todos",
            expand=1, height=55,
        )
        self.dd_furgon.on_change = lambda _: self._actualizar_preview()
        self.in_titulo = ft.TextField(
            label="Título personalizado (opcional)",
            expand=2, height=55,
        )

        _puede_exportar = estado.puede_hacer("graficos", "exportar_pdf")
        self.btn_generar  = ft.ElevatedButton("VER GRÁFICO",          icon=ft.Icons.BAR_CHART,      bgcolor="blue800",  color="white", height=45, on_click=lambda _: self._actualizar_preview(), tooltip="Generar y mostrar el gráfico seleccionado")
        self.btn_pdf      = ft.ElevatedButton("EXPORTAR PDF",         icon=ft.Icons.PICTURE_AS_PDF, bgcolor="red700",   color="white", height=45, on_click=self.accion_exportar_pdf,        visible=_puede_exportar, tooltip="Exportar gráfico actual como PDF")
        self.btn_excel    = ft.ElevatedButton("EXPORTAR EXCEL",       icon=ft.Icons.TABLE_CHART,    bgcolor="green700", color="white", height=45, on_click=self.accion_exportar_excel,      visible=_puede_exportar, tooltip="Exportar datos como archivo Excel")
        self.btn_informe  = ft.ElevatedButton("INFORME COMPLETO PDF", icon=ft.Icons.SUMMARIZE,      bgcolor="teal700",  color="white", height=45, on_click=self.accion_informe_completo,    visible=_puede_exportar, tooltip="Generar informe PDF con todos los gráficos")

        self.txt_estado      = ft.Text("Carga los datos con el botón actualizar, luego selecciona un gráfico.", size=12, color="grey600", italic=True)
        self.txt_descripcion = ft.Text("", size=12, color="blue900", italic=True, visible=False)
        self.progress        = ft.ProgressBar(visible=False, color="blue700")

        self.img_grafico = ft.Image(src="", visible=False, fit="contain", expand=True)
        self.cont_grafico = ft.Container(
            content=self.img_grafico,
            visible=False,
            expand=True,
            bgcolor="white",
            border_radius=10,
            border=ft.border.all(1, "grey200"),
            padding=10,
            margin=ft.margin.only(top=10),
        )

        from utils.ui_helpers import sandwich

        _header = ft.Column([
            ft.Row([
                ft.Text("Gráficos y Análisis", size=22, weight="bold", color="blue900", expand=True),
                ft.IconButton(ft.Icons.REFRESH, tooltip="Recargar datos", on_click=lambda _: self.inicializar()),
            ]),
            ft.Divider(height=6, color="transparent"),
            ft.Container(
                content=ft.Column([
                    ft.Row([self.dd_tipo, self.dd_familia, self.in_titulo], spacing=12),
                    ft.Row([self.dd_bodega, self.dd_furgon], spacing=12),
                    self.txt_descripcion,
                    ft.Divider(height=4, color="transparent"),
                    self.progress,
                    self.txt_estado,
                ]),
                padding=20, bgcolor="white", border_radius=10, border=ft.border.all(1, "grey300"),
            ),
        ], spacing=0)

        _footer = ft.Row([self.btn_generar, self.btn_pdf, self.btn_excel, self.btn_informe], spacing=10)

        # cont_grafico ya tiene expand=True; lo usamos directamente como body
        self.cont_grafico.scroll = ft.ScrollMode.AUTO

        self.content = sandwich(_header, self.cont_grafico, _footer)

    # ══════════════════════════════ Cache ══════════════════════════════
    def pre_cargar_si_cache_listo(self):
        from core.cache import app_cache
        if getattr(self, '_datos_cargados', False) or not app_cache.gra_ready.is_set():
            return
        self._poblar_desde_cache(app_cache, llamar_update=False)

    def _poblar_desde_cache(self, cache, llamar_update=True):
        self.datos_productos   = list(cache.gra_productos   or [])
        self.datos_auditoria   = list(cache.gra_auditoria   or [])
        self.datos_movimientos = list(cache.gra_movimientos or [])
        self.datos_bodegas     = list(cache.gra_bodegas     or [])
        self.datos_bp          = list(cache.gra_bp          or [])
        self.datos_furgones    = list(cache.gra_furgones    or [])
        self.datos_fp          = list(cache.gra_fp          or [])

        fams = sorted(set(p["familia"] for p in self.datos_productos if p.get("familia")))
        self.familias_unicas = ["Todas"] + fams
        self.dd_familia.options = [ft.dropdown.Option(f) for f in self.familias_unicas]
        self.dd_bodega.options  = [ft.dropdown.Option("Todas")] + [ft.dropdown.Option(b["nombre"]) for b in self.datos_bodegas]
        self.dd_furgon.options  = [ft.dropdown.Option("Todos")] + [ft.dropdown.Option(f["nombre"]) for f in self.datos_furgones]

        n = len(self.datos_productos)
        self.txt_estado.value = f"Datos cargados: {n} productos. Listo para exportar."
        self._datos_cargados = True
        print(f"[GRAFICOS] pre-cargados ({n} productos)")

        if llamar_update and self.page_ref:
            self.dd_familia.update()
            self.dd_bodega.update()
            self.dd_furgon.update()
            self.txt_estado.update()

    # ══════════════════════════════ Carga de datos ══════════════════════════════
    def inicializar(self):
        from core.cache import app_cache
        if getattr(self, '_datos_cargados', False):
            if self.page_ref:
                try:
                    self.dd_familia.update()
                    self.dd_bodega.update()
                    self.dd_furgon.update()
                except Exception: pass
            return
        # Si el cache ya está listo, poblar directamente
        if app_cache.gra_ready.is_set():
            self._poblar_desde_cache(app_cache, llamar_update=True)
            return

        def _run():
            try:
                self.progress.visible = True
                self.txt_estado.value = "Descargando datos..."
                if self.page_ref:
                    self.progress.update()
                    self.txt_estado.update()

                # Esperar cache (máx 60s); si llega, usarlo; si no, fetch propio
                app_cache.gra_ready.wait(timeout=60)
                if app_cache.gra_productos:
                    self._poblar_desde_cache(app_cache, llamar_update=True)
                    self.progress.visible = False
                    if self.page_ref:
                        self.progress.update()
                        try: self.page_ref.update()
                        except Exception: pass
                    return

                sb = get_sb()
                res_p = sb.table("productos").select(
                    "sku, nombre, stock_global, costo_neto, precio_venta, familias(nombre), subfamilias(nombre)"
                ).limit(15000).execute()
                for p in res_p.data or []:
                    p["stock_actual"] = p.get("stock_global", 0) or 0
                    p["familia"]    = (p.get("familias")    or {}).get("nombre") or "Sin Familia"
                    p["subfamilia"] = (p.get("subfamilias") or {}).get("nombre") or "Sin Subfamilia"
                self.datos_productos = res_p.data or []
                self.datos_auditoria   = sb.table("auditoria").select("usuario, accion, fecha").limit(5000).execute().data or []
                self.datos_movimientos = sb.table("movimientos_inventario").select("sku, tipo_movimiento, cantidad, fecha, usuario").limit(10000).execute().data or []
                self.datos_bodegas = sb.table("bodegas").select("id, nombre").order("id").execute().data or []
                self.datos_bp = fetch_all("bodega_productos", "bodega_id, sku, cantidad")
                self.datos_furgones = fetch_furgones_con_totales()
                self.datos_fp = fetch_all("furgon_productos", "furgon_id, sku, cantidad")

                fams = sorted(set(p["familia"] for p in self.datos_productos if p.get("familia")))
                self.familias_unicas = ["Todas"] + fams
                self.dd_familia.options = [ft.dropdown.Option(f) for f in self.familias_unicas]
                self.dd_bodega.options  = [ft.dropdown.Option("Todas")] + [ft.dropdown.Option(b["nombre"]) for b in self.datos_bodegas]
                self.dd_furgon.options  = [ft.dropdown.Option("Todos")] + [ft.dropdown.Option(f["nombre"]) for f in self.datos_furgones]

                self.progress.visible = False
                self.txt_estado.value = f"Datos cargados: {len(self.datos_productos)} productos. Listo para exportar."
                self._datos_cargados = True
                if self.page_ref:
                    self.dd_familia.update()
                    self.dd_bodega.update()
                    self.dd_furgon.update()
                    self.progress.update()
                    self.txt_estado.update()
                    try: self.page_ref.update()
                    except Exception: pass
            except Exception as ex:
                self.progress.visible = False
                self.txt_estado.value = f"Error al cargar: {ex}"
                if self.page_ref:
                    self.progress.update()
                    self.txt_estado.update()
        hilo(_run)

    def _actualizar_preview(self):
        tipo = self.dd_tipo.value or ""
        if tipo.startswith("__sep__") or not tipo:
            return

        desc = _DESCRIPCIONES.get(tipo, "")
        self.txt_descripcion.value = desc
        self.txt_descripcion.visible = bool(desc)
        if self.page_ref:
            self.txt_descripcion.update()

        if not self.datos_productos and tipo not in (
            "movimientos_tipo","movimientos_mes","top_despachos","top_usuarios","acciones_tipo"
        ):
            self.txt_estado.value = "Primero recarga los datos con el botón actualizar."
            if self.page_ref: self.txt_estado.update()
            return

        titulo = (self.in_titulo.value or "").strip() or _TIPO_LABEL.get(tipo, tipo).strip()
        datos = self._filtrar_productos()

        def _run():
            self.progress.visible = True
            self.txt_estado.value = "Generando gráfico..."
            if self.page_ref:
                self.progress.update()
                self.txt_estado.update()
            try:
                ruta = self._generar_png_temporal(tipo, titulo, datos)
                self.img_grafico.src = ruta
                self.img_grafico.visible = True
                self.cont_grafico.visible = True
                self.txt_estado.value = _TIPO_LABEL.get(tipo, tipo)
            except Exception as ex:
                self.txt_estado.value = f"Error al generar gráfico: {ex}"
            self.progress.visible = False
            if self.page_ref:
                self.img_grafico.update()
                self.cont_grafico.update()
                self.progress.update()
                self.txt_estado.update()
        hilo(_run)

    # ══════════════════════════════ Motor de gráficos ══════════════════════════════
    def _filtrar_productos(self):
        fam = self.dd_familia.value or "Todas"
        bod = self.dd_bodega.value or "Todas"

        prods = self.datos_productos if fam == "Todas" else [
            p for p in self.datos_productos if p.get("familia") == fam
        ]

        if bod != "Todas":
            bod_id = next((b["id"] for b in self.datos_bodegas if b["nombre"] == bod), None)
            if bod_id is not None:
                bp_map = {r["sku"]: r.get("cantidad", 0) or 0
                          for r in self.datos_bp if r["bodega_id"] == bod_id}
                prods = [{**p, "stock_actual": bp_map[p["sku"]]}
                         for p in prods if p["sku"] in bp_map]

        return prods

    def _renderizar(self, tipo: str, titulo: str, datos_prod, fig=None, ax=None):
        """Dibuja el gráfico pedido. Devuelve (fig, ax) o los recibidos."""
        externo = fig is not None
        if not externo:
            if tipo in _FIGSIZE_PIE:
                fig, ax = plt.subplots(figsize=(7, 7))
            else:
                fig, ax = plt.subplots(figsize=(11, 5.5))

        colores = ["#1565C0","#2E7D32","#6A1B9A","#E65100","#00838F",
                   "#AD1457","#F57F17","#37474F","#558B2F","#4527A0"]

        # ── 1. Stock por familia ──────────────────────────────────────
        if tipo == "stock_familia":
            ag = defaultdict(int)
            for p in datos_prod:
                ag[p.get("familia","Sin Familia")] += p.get("stock_actual",0)
            ag = dict(sorted(ag.items(), key=lambda x: x[1], reverse=True)[:15])
            bars = ax.bar(ag.keys(), ag.values(), color=colores[:len(ag)])
            ax.bar_label(bars, padding=3, fontsize=8)
            plt.xticks(rotation=40, ha="right", fontsize=8)
            ax.set_ylabel("Unidades")

        # ── 2. Stock por subfamilia ───────────────────────────────────
        elif tipo == "stock_subfamilia":
            ag = defaultdict(int)
            for p in datos_prod:
                ag[p.get("subfamilia","Sin Subfamilia")] += p.get("stock_actual",0)
            ag = dict(sorted(ag.items(), key=lambda x: x[1], reverse=True)[:15])
            bars = ax.barh(list(ag.keys()), list(ag.values()), color="#1565C0")
            ax.bar_label(bars, padding=3, fontsize=8)
            ax.invert_yaxis()
            ax.set_xlabel("Unidades")

        # ── 3. Stock por bodega ───────────────────────────────────────
        elif tipo == "stock_bodega":
            bod_sel = self.dd_bodega.value or "Todas"
            bod_map = {b["id"]: b["nombre"] for b in self.datos_bodegas}
            if bod_sel != "Todas":
                bod_id = next((b["id"] for b in self.datos_bodegas if b["nombre"] == bod_sel), None)
                prod_map = {p["sku"]: p.get("nombre", p["sku"]) for p in self.datos_productos}
                items = sorted(
                    [(prod_map.get(r["sku"], r["sku"]), r.get("cantidad", 0) or 0)
                     for r in self.datos_bp if r["bodega_id"] == bod_id],
                    key=lambda x: x[1], reverse=True
                )[:15] if bod_id else []
                if items:
                    labs, vals = zip(*items)
                    y = np.arange(len(labs))
                    bars = ax.barh(y, vals, color="#1565C0")
                    ax.bar_label(bars, padding=3, fontsize=8)
                    ax.set_yticks(y); ax.set_yticklabels([l[:30] for l in labs], fontsize=8)
                    ax.invert_yaxis(); ax.set_xlabel("Unidades")
                else:
                    ax.text(0.5, 0.5, "Sin productos en esta bodega", ha="center", va="center", color="grey")
            else:
                ag = defaultdict(int)
                for r in self.datos_bp:
                    ag[bod_map.get(r["bodega_id"], f"Bodega {r['bodega_id']}")] += r.get("cantidad", 0) or 0
                if not ag:
                    ax.text(0.5, 0.5, "Sin datos de bodegas", ha="center", va="center", color="grey")
                else:
                    bars = ax.bar(ag.keys(), ag.values(), color=colores[:len(ag)])
                    ax.bar_label(bars, padding=3)
                    plt.xticks(rotation=20, ha="right")
                ax.set_ylabel("Unidades totales")

        # ── 4. Salud del stock ────────────────────────────────────────
        elif tipo == "salud_stock":
            sano    = sum(1 for p in datos_prod if (p.get("stock_actual") or 0) > 3)
            critico = sum(1 for p in datos_prod if 0 < (p.get("stock_actual") or 0) <= 3)
            cero    = sum(1 for p in datos_prod if (p.get("stock_actual") or 0) == 0)
            labels  = [f"Sano (>3)\n{sano}", f"Crítico (1-3)\n{critico}", f"Sin stock\n{cero}"]
            colors  = ["#2E7D32", "#EF6C00", "#C62828"]
            ax.pie([sano, critico, cero], labels=labels, autopct="%1.1f%%",
                   colors=colors, wedgeprops=dict(width=0.45, edgecolor="w"))
            ax.text(0, 0, str(len(datos_prod)), ha="center", va="center",
                    fontsize=14, fontweight="bold", color="#1A237E")

        # ── 5. Top 10 mayor stock ─────────────────────────────────────
        elif tipo == "top_stock":
            top = sorted(datos_prod, key=lambda p: p.get("stock_actual",0), reverse=True)[:10]
            labs = [p.get("nombre","")[:28] for p in top]
            vals = [p.get("stock_actual",0) for p in top]
            y = np.arange(len(labs))
            bars = ax.barh(y, vals, color="#0277BD")
            ax.bar_label(bars, padding=3, fontsize=8)
            ax.set_yticks(y); ax.set_yticklabels(labs, fontsize=8)
            ax.invert_yaxis(); ax.set_xlabel("Stock")

        # ── 6. Familias con más productos sin stock ───────────────────
        elif tipo == "top_sin_stock":
            ag = defaultdict(int)
            for p in datos_prod:
                if (p.get("stock_actual") or 0) == 0:
                    ag[p.get("familia","Sin Familia")] += 1
            ag = dict(sorted(ag.items(), key=lambda x: x[1], reverse=True)[:12])
            if not ag:
                ax.text(0.5, 0.5, "Todos los productos tienen stock", ha="center", va="center", color="grey")
            else:
                bars = ax.bar(ag.keys(), ag.values(), color="#B71C1C")
                ax.bar_label(bars, padding=3)
                plt.xticks(rotation=40, ha="right", fontsize=8)
            ax.set_ylabel("Cantidad de productos sin stock")

        # ── 7. Valor inventario por familia ──────────────────────────
        elif tipo == "valor_familia":
            fc, fv = defaultdict(float), defaultdict(float)
            for p in datos_prod:
                f = p.get("familia","Sin Familia")
                s = p.get("stock_actual",0)
                fc[f] += s * (p.get("costo_neto") or 0)
                fv[f] += s * (p.get("precio_venta") or 0)
            fams = [f for f, _ in sorted(fc.items(), key=lambda x: x[1], reverse=True)[:10]]
            x = np.arange(len(fams)); w = 0.35
            ax.bar(x - w/2, [fc[f]/1e6 for f in fams], w, label="Costo", color="#1976D2")
            ax.bar(x + w/2, [fv[f]/1e6 for f in fams], w, label="Venta", color="#2E7D32")
            ax.set_xticks(x); ax.set_xticklabels(fams, rotation=40, ha="right", fontsize=8)
            ax.set_ylabel("Millones ($)"); ax.legend()

        # ── 8. Valor inventario por bodega ────────────────────────────
        elif tipo == "valor_bodega":
            bod_sel  = self.dd_bodega.value or "Todas"
            bod_map  = {b["id"]: b["nombre"] for b in self.datos_bodegas}
            prod_map = {p["sku"]: p for p in self.datos_productos}
            if bod_sel != "Todas":
                bod_id = next((b["id"] for b in self.datos_bodegas if b["nombre"] == bod_sel), None)
                items = []
                for r in self.datos_bp:
                    if r["bodega_id"] != bod_id: continue
                    p = prod_map.get(r["sku"], {})
                    cant = r.get("cantidad", 0) or 0
                    venta = cant * (p.get("precio_venta") or 0)
                    if venta > 0:
                        items.append((p.get("nombre", r["sku"])[:28], cant * (p.get("costo_neto") or 0), venta))
                items = sorted(items, key=lambda x: x[2], reverse=True)[:12]
                if items:
                    labs = [i[0] for i in items]
                    cs   = [i[1]/1e3 for i in items]
                    vs   = [i[2]/1e3 for i in items]
                    y = np.arange(len(labs)); w = 0.35
                    ax.barh(y - w/2, cs, w, label="Costo (miles $)", color="#1976D2")
                    ax.barh(y + w/2, vs, w, label="Venta (miles $)", color="#2E7D32")
                    ax.set_yticks(y); ax.set_yticklabels(labs, fontsize=8)
                    ax.invert_yaxis(); ax.set_xlabel("Miles ($)"); ax.legend()
                else:
                    ax.text(0.5, 0.5, "Sin datos en esta bodega", ha="center", va="center", color="grey")
            else:
                bc, bv = defaultdict(float), defaultdict(float)
                for r in self.datos_bp:
                    p = prod_map.get(r["sku"], {})
                    nombre = bod_map.get(r["bodega_id"], f"Bodega {r['bodega_id']}")
                    cant = r.get("cantidad", 0) or 0
                    bc[nombre] += cant * (p.get("costo_neto") or 0)
                    bv[nombre] += cant * (p.get("precio_venta") or 0)
                if not bc:
                    ax.text(0.5, 0.5, "Sin datos de bodegas", ha="center", va="center", color="grey")
                else:
                    bods = list(bc.keys())
                    x = np.arange(len(bods)); w = 0.35
                    ax.bar(x - w/2, [bc[b]/1e6 for b in bods], w, label="Costo", color="#1976D2")
                    ax.bar(x + w/2, [bv[b]/1e6 for b in bods], w, label="Venta", color="#2E7D32")
                    ax.set_xticks(x); ax.set_xticklabels(bods, rotation=20, ha="right")
                    ax.set_ylabel("Millones ($)"); ax.legend()

        # ── 9. Margen promedio por familia (%) ────────────────────────
        elif tipo == "margen_familia":
            ag = defaultdict(list)
            for p in datos_prod:
                v = p.get("precio_venta") or 0
                c = p.get("costo_neto") or 0
                if v > 0:
                    ag[p.get("familia","Sin Familia")].append((v - c) / v * 100)
            avg = {f: sum(m)/len(m) for f, m in ag.items()}
            avg = dict(sorted(avg.items(), key=lambda x: x[1], reverse=True)[:15])
            bars = ax.bar(avg.keys(), avg.values(), color="#2E7D32")
            ax.bar_label(bars, fmt="%.1f%%", padding=3, fontsize=8)
            plt.xticks(rotation=40, ha="right", fontsize=8)
            ax.set_ylabel("Margen promedio (%)")
            ax.axhline(0, color="grey", linewidth=0.8, linestyle="--")

        # ── 10. Top 10 productos por margen absoluto ($) ─────────────
        elif tipo == "top_margen_prod":
            top = sorted(
                [p for p in datos_prod if (p.get("precio_venta") or 0) > 0],
                key=lambda p: (p.get("precio_venta",0) or 0) - (p.get("costo_neto",0) or 0),
                reverse=True
            )[:10]
            labs = [p.get("nombre","")[:28] for p in top]
            vals = [(p.get("precio_venta",0) or 0) - (p.get("costo_neto",0) or 0) for p in top]
            y = np.arange(len(labs))
            bars = ax.barh(y, vals, color="#2E7D32")
            ax.bar_label(bars, fmt="$%.0f", padding=3, fontsize=8)
            ax.set_yticks(y); ax.set_yticklabels(labs, fontsize=8)
            ax.invert_yaxis(); ax.set_xlabel("Margen por unidad ($)")

        # ── 11. Anomalías financieras ─────────────────────────────────
        elif tipo == "anomalias":
            sin_precio = sum(1 for p in datos_prod if not (p.get("precio_venta") or 0))
            sin_costo  = sum(1 for p in datos_prod if not (p.get("costo_neto") or 0))
            perdida    = sum(1 for p in datos_prod
                            if (p.get("costo_neto") or 0) > (p.get("precio_venta") or 0) > 0)
            bars = ax.bar(["Sin precio", "Sin costo", "Venta a pérdida"],
                          [sin_precio, sin_costo, perdida],
                          color=["#E65100","#F57F17","#B71C1C"])
            ax.bar_label(bars, padding=3)
            ax.set_ylabel("Cantidad de productos")

        # ── 12. Dispersión costo vs venta ─────────────────────────────
        elif tipo == "dispersion":
            cs = [(p.get("costo_neto") or 0) for p in datos_prod]
            vs = [(p.get("precio_venta") or 0) for p in datos_prod]
            col_pts = ["green" if v > c else "red" if c > v else "orange" for c, v in zip(cs, vs)]
            ax.scatter(cs, vs, alpha=0.4, c=col_pts, s=18)
            mx = max(max(cs+[0]), max(vs+[0]))
            ax.plot([0, mx], [0, mx], "r--", alpha=0.3, label="Equilibrio")
            ax.set_xlabel("Costo Neto ($)"); ax.set_ylabel("Precio Venta ($)")
            ax.legend()

        # ── 13. Movimientos por tipo ──────────────────────────────────
        elif tipo == "movimientos_tipo":
            ag = Counter(m.get("tipo_movimiento","") for m in self.datos_movimientos)
            if not ag:
                ax.text(0.5, 0.5, "Sin movimientos registrados", ha="center", va="center", color="grey")
            else:
                labels_limpios = [k.replace("_"," ").title() for k in ag.keys()]
                ax.pie(ag.values(), labels=labels_limpios, autopct="%1.1f%%",
                       colors=colores[:len(ag)], startangle=140)
                ax.axis("equal")

        # ── 14. Evolución mensual de movimientos ──────────────────────
        elif tipo == "movimientos_mes":
            por_mes = defaultdict(int)
            for m in self.datos_movimientos:
                fecha = m.get("fecha","")
                if fecha and len(fecha) >= 7:
                    por_mes[fecha[:7]] += 1
            if not por_mes:
                ax.text(0.5, 0.5, "Sin movimientos registrados", ha="center", va="center", color="grey")
            else:
                meses = sorted(por_mes.keys())[-18:]
                vals  = [por_mes[m] for m in meses]
                ax.plot(meses, vals, marker="o", color="#1565C0", linewidth=2)
                ax.fill_between(meses, vals, alpha=0.15, color="#1565C0")
                plt.xticks(rotation=45, ha="right", fontsize=8)
                ax.set_ylabel("Cantidad de movimientos")
                ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))

        # ── 15. Top 10 productos más despachados a terreno ────────────
        elif tipo == "top_despachos":
            ag = defaultdict(int)
            for m in self.datos_movimientos:
                if m.get("tipo_movimiento") == "DESPACHO_TECNICO":
                    ag[m.get("sku","")] += m.get("cantidad",0) or 0
            top = sorted(ag.items(), key=lambda x: x[1], reverse=True)[:10]
            if not top:
                ax.text(0.5, 0.5, "Sin despachos técnicos registrados", ha="center", va="center", color="grey")
            else:
                prod_map = {p["sku"]: p.get("nombre","") for p in self.datos_productos}
                labs = [prod_map.get(s, s)[:28] for s, _ in top]
                vals = [v for _, v in top]
                y = np.arange(len(labs))
                bars = ax.barh(y, vals, color="#6A1B9A")
                ax.bar_label(bars, padding=3, fontsize=8)
                ax.set_yticks(y); ax.set_yticklabels(labs, fontsize=8)
                ax.invert_yaxis(); ax.set_xlabel("Unidades despachadas")

        # ── 17. Distribución por familia (pie) ────────────────────────
        elif tipo == "dist_familia":
            cnt = Counter(p.get("familia","Sin Familia") for p in datos_prod)
            top = dict(cnt.most_common(10))
            ax.pie(top.values(), labels=top.keys(), autopct="%1.1f%%",
                   startangle=140, colors=plt.cm.Paired.colors)
            ax.axis("equal")

        # ── 18. Distribución por subfamilia (pie) ─────────────────────
        elif tipo == "dist_subfamilia":
            cnt = Counter(p.get("subfamilia","Sin Subfamilia") for p in datos_prod)
            top = dict(cnt.most_common(12))
            ax.pie(top.values(), labels=top.keys(), autopct="%1.1f%%",
                   startangle=140, colors=plt.cm.Set3.colors)
            ax.axis("equal")

        # ── 19. Comparativa bodegas vs furgones ──────────────────────
        elif tipo == "comparativa_bd_fg":
            total_bd = sum(r.get("cantidad",0) or 0 for r in self.datos_bp)
            total_fg = sum(r.get("cantidad",0) or 0 for r in self.datos_fp)
            bars = ax.bar(["Bodegas", "Furgones"], [total_bd, total_fg],
                          color=["#1565C0","#4527A0"], width=0.4)
            ax.bar_label(bars, padding=4, fontsize=10)
            ax.set_ylabel("Unidades totales")

        # ── 20. Unidades por furgón ───────────────────────────────────
        elif tipo == "furgon_stock":
            furg_sel = self.dd_furgon.value or "Todos"
            if furg_sel != "Todos":
                furg_id = next((f["id"] for f in self.datos_furgones if f["nombre"] == furg_sel), None)
                prod_map = {p["sku"]: p.get("nombre", p["sku"]) for p in self.datos_productos}
                items = sorted(
                    [(prod_map.get(r["sku"], r["sku"]), r.get("cantidad", 0) or 0)
                     for r in self.datos_fp if r["furgon_id"] == furg_id],
                    key=lambda x: x[1], reverse=True
                )[:15] if furg_id else []
                if items:
                    labs, vals = zip(*items)
                    y = np.arange(len(labs))
                    bars = ax.barh(y, vals, color="#4527A0")
                    ax.bar_label(bars, padding=3, fontsize=8)
                    ax.set_yticks(y); ax.set_yticklabels([l[:30] for l in labs], fontsize=8)
                    ax.invert_yaxis(); ax.set_xlabel("Unidades")
                else:
                    ax.text(0.5, 0.5, "Sin productos en este furgón", ha="center", va="center", color="grey")
            else:
                labs = [f["nombre"] for f in self.datos_furgones]
                vals = [f.get("_total_uds", 0) for f in self.datos_furgones]
                if not labs:
                    ax.text(0.5, 0.5, "Sin furgones registrados", ha="center", va="center", color="grey")
                else:
                    bars = ax.bar(labs, vals, color="#4527A0")
                    ax.bar_label(bars, padding=3)
                    plt.xticks(rotation=20, ha="right")
                ax.set_ylabel("Unidades")

        # ── 21. Valor por furgón ──────────────────────────────────────
        elif tipo == "furgon_valor":
            furg_sel = self.dd_furgon.value or "Todos"
            furg_map = {f["id"]: f["nombre"] for f in self.datos_furgones}
            prod_map = {p["sku"]: p for p in self.datos_productos}
            if furg_sel != "Todos":
                furg_id = next((f["id"] for f in self.datos_furgones if f["nombre"] == furg_sel), None)
                items = []
                for r in self.datos_fp:
                    if r["furgon_id"] != furg_id: continue
                    p = prod_map.get(r["sku"], {})
                    cant = r.get("cantidad", 0) or 0
                    venta = cant * (p.get("precio_venta") or 0)
                    if venta > 0:
                        items.append((p.get("nombre", r["sku"])[:28], cant * (p.get("costo_neto") or 0), venta))
                items = sorted(items, key=lambda x: x[2], reverse=True)[:12]
                if items:
                    labs = [i[0] for i in items]
                    cs   = [i[1]/1e3 for i in items]
                    vs   = [i[2]/1e3 for i in items]
                    y = np.arange(len(labs)); w = 0.35
                    ax.barh(y - w/2, cs, w, label="Costo (miles $)", color="#1976D2")
                    ax.barh(y + w/2, vs, w, label="Venta (miles $)", color="#2E7D32")
                    ax.set_yticks(y); ax.set_yticklabels(labs, fontsize=8)
                    ax.invert_yaxis(); ax.set_xlabel("Miles ($)"); ax.legend()
                else:
                    ax.text(0.5, 0.5, "Sin datos en este furgón", ha="center", va="center", color="grey")
            else:
                fc, fv = defaultdict(float), defaultdict(float)
                for r in self.datos_fp:
                    p = prod_map.get(r["sku"], {})
                    nombre = furg_map.get(r["furgon_id"], f"Furgón {r['furgon_id']}")
                    cant = r.get("cantidad", 0) or 0
                    fc[nombre] += cant * (p.get("costo_neto") or 0)
                    fv[nombre] += cant * (p.get("precio_venta") or 0)
                if not fc:
                    ax.text(0.5, 0.5, "Sin furgones con stock", ha="center", va="center", color="grey")
                else:
                    furgs = list(fc.keys())
                    x = np.arange(len(furgs)); w = 0.35
                    ax.bar(x - w/2, [fc[f]/1e6 for f in furgs], w, label="Costo", color="#1976D2")
                    ax.bar(x + w/2, [fv[f]/1e6 for f in furgs], w, label="Venta", color="#2E7D32")
                    ax.set_xticks(x); ax.set_xticklabels(furgs, rotation=20, ha="right")
                    ax.set_ylabel("Millones ($)"); ax.legend()

        # ── 22. Top 10 usuarios activos ───────────────────────────────
        elif tipo == "top_usuarios":
            cnt = Counter(a.get("usuario") for a in self.datos_auditoria if a.get("usuario"))
            top = dict(cnt.most_common(10))
            if not top:
                ax.text(0.5, 0.5, "Sin registros de auditoría", ha="center", va="center", color="grey")
            else:
                y = np.arange(len(top))
                bars = ax.barh(y, list(top.values()), color="#00838F")
                ax.bar_label(bars, padding=3, fontsize=8)
                ax.set_yticks(y); ax.set_yticklabels(list(top.keys()), fontsize=8)
                ax.invert_yaxis(); ax.set_xlabel("Acciones registradas")

        # ── 23. Distribución de acciones en auditoría ─────────────────
        elif tipo == "acciones_tipo":
            cnt = Counter(a.get("accion","") for a in self.datos_auditoria)
            top = dict(cnt.most_common(10))
            if not top:
                ax.text(0.5, 0.5, "Sin registros de auditoría", ha="center", va="center", color="grey")
            else:
                y = np.arange(len(top))
                bars = ax.barh(y, list(top.values()), color="#AD1457")
                ax.bar_label(bars, padding=3, fontsize=8)
                ax.set_yticks(y); ax.set_yticklabels(list(top.keys()), fontsize=8)
                ax.invert_yaxis(); ax.set_xlabel("Veces registrada")

        else:
            ax.text(0.5, 0.5, f"Gráfico '{tipo}' no implementado",
                    ha="center", va="center", color="grey")

        ax.set_title(titulo, pad=16, fontweight="bold", color="#1A237E", fontsize=12)
        plt.tight_layout()
        return fig, ax

    # ══════════════════════════════ Acciones ══════════════════════════════
    def _generar_png_temporal(self, tipo, titulo, datos):
        """Genera el gráfico y lo guarda en un archivo temporal. Devuelve la ruta."""
        import tempfile
        fig, _ = self._renderizar(tipo, titulo, datos)
        fd, ruta = tempfile.mkstemp(suffix=".png")
        os.close(fd)
        fig.savefig(ruta, dpi=130, bbox_inches="tight")
        plt.close(fig)
        return ruta

    def accion_exportar_pdf(self, e):
        tipo = self.dd_tipo.value or "stock_familia"
        if tipo.startswith("__sep__"):
            self.mostrar_snack("Selecciona un tipo de gráfico válido.", "orange"); return
        if not self.datos_productos and tipo not in ("movimientos_tipo","movimientos_mes","top_despachos","top_usuarios","acciones_tipo"):
            self.mostrar_snack("Carga los datos primero (botón actualizar).", "orange"); return

        titulo = (self.in_titulo.value or "").strip() or _TIPO_LABEL.get(tipo, tipo).strip()
        datos  = self._filtrar_productos()

        def _run():
            import base64 as _b64
            try: from fpdf import FPDF
            except ImportError:
                self.mostrar_snack("Falta fpdf: pip install fpdf", "red"); return

            self.progress.visible = True
            self.txt_estado.value = "Generando gráfico..."
            if self.page_ref:
                self.progress.update()
                self.txt_estado.update()

            try:
                ruta_img = self._generar_png_temporal(tipo, titulo, datos)
                pdf = FPDF(orientation="L", unit="mm", format="A4")
                pdf.add_page()
                pdf.set_font("Arial","B",14)
                pdf.cell(0, 10, titulo.encode("latin-1","replace").decode("latin-1"), ln=True, align="C")
                if tipo in _FIGSIZE_PIE:
                    pdf.image(ruta_img, x=68, y=26, w=160)
                else:
                    pdf.image(ruta_img, x=10, y=28, w=277)
                import subprocess as _sp, datetime as _dt
                _fname = f"/tmp/Grafico_{tipo}_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                open(_fname, 'wb').write(pdf.output())
                _sp.Popen(['open', _fname])
                os.unlink(ruta_img)
                self.txt_estado.value = "PDF generado."
                self.mostrar_snack("PDF generado.", "green700")
            except Exception as ex:
                self.mostrar_snack(f"Error PDF: {ex}", "red")
            self.progress.visible = False
            if self.page_ref:
                self.progress.update()
                self.txt_estado.update()
        hilo(_run)

    def accion_exportar_excel(self, e):
        if not self.datos_productos:
            self.mostrar_snack("Sin datos para exportar.", "orange"); return

        def _run():
            import openpyxl, io, base64 as _b64
            from openpyxl.styles import Font, PatternFill
            try:
                datos = self._filtrar_productos()
                ver_costo = estado.puede_ver("costo")
                wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Productos"
                hdrs = ["SKU","Nombre","Familia","Subfamilia","Stock","Precio Venta ($)"]
                if ver_costo:
                    hdrs.insert(5, "Costo ($)")
                ws.append(hdrs)
                fill = PatternFill("solid", fgColor="1565C0")
                for ci in range(1, len(hdrs)+1):
                    c = ws.cell(1, ci)
                    c.fill = fill; c.font = Font(bold=True, color="FFFFFF")
                for p in datos:
                    fila = [p.get("sku",""), p.get("nombre",""), p.get("familia",""),
                            p.get("subfamilia",""), p.get("stock_actual",0)]
                    if ver_costo:
                        fila.append(p.get("costo_neto",0))
                    fila.append(p.get("precio_venta",0))
                    ws.append(fila)
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = min(
                        max(len(str(c.value or "")) for c in col) + 4, 45)
                _buf = io.BytesIO(); wb.save(_buf)
                import subprocess as _sp, datetime as _dt
                _fname = f"/tmp/Grafico_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                open(_fname, 'wb').write(_buf.getvalue())
                _sp.Popen(['open', _fname])
                self.mostrar_snack("Excel generado.", "green700")
            except Exception as ex:
                self.mostrar_snack(f"Error Excel: {ex}", "red")
        hilo(_run)

    def accion_informe_completo(self, e):
        if not self.datos_productos:
            self.mostrar_snack("Sin datos para el informe.", "orange"); return

        def _run():
            import base64 as _b64
            try: from fpdf import FPDF
            except ImportError:
                self.mostrar_snack("Falta fpdf: pip install fpdf", "red"); return

            try:
                self.progress.visible = True
                datos = self.datos_productos
                tipos_informe = [k for k, _ in TIPOS_GRAFICOS]
                total = len(tipos_informe)

                pdf = FPDF(orientation="L", unit="mm", format="A4")

                for i, tipo in enumerate(tipos_informe):
                    self.txt_estado.value = f"Generando página {i+1}/{total}: {_TIPO_LABEL.get(tipo,tipo)}"
                    if self.page_ref: self.txt_estado.update()

                    titulo = _TIPO_LABEL.get(tipo, tipo).strip()
                    try:
                        fig, _ = self._renderizar(tipo, titulo, datos)
                        ruta_img = os.path.join(_ASSETS_DIR, f"_informe_{tipo}.png")
                        fig.savefig(ruta_img, dpi=110, bbox_inches="tight")
                        plt.close(fig)
                        pdf.add_page()
                        pdf.set_font("Arial","B",13)
                        pdf.cell(0, 9, titulo.encode("latin-1","replace").decode("latin-1"), ln=True, align="C")
                        if tipo in _FIGSIZE_PIE:
                            pdf.image(ruta_img, x=68, y=26, w=160)
                        else:
                            pdf.image(ruta_img, x=10, y=26, w=277)
                    except Exception:
                        pass

                import subprocess as _sp, datetime as _dt
                _fname = f"/tmp/InformeCompleto_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                open(_fname, 'wb').write(pdf.output())
                _sp.Popen(['open', _fname])
                self.progress.visible = False
                self.txt_estado.value = f"Informe completo ({total} páginas) exportado."
                self.mostrar_snack(f"Informe de {total} gráficos generado.", "teal700")
            except Exception as ex:
                self.progress.visible = False
                self.txt_estado.value = f"Error: {ex}"
                self.mostrar_snack(f"Error: {ex}", "red")
            if self.page_ref:
                self.progress.update()
                self.txt_estado.update()
        hilo(_run)
