# vistas/movimientos.py
import flet as ft
from core.database import get_sb, ejecutar_traspaso, cache_invalidate
from core.estado import estado
from core.utilidades import hilo, registrar_auditoria
from core.widgets import BuscadorProducto


def _label_bodega(b: dict) -> str:
    area = (b.get("areas") or {}).get("nombre") or ""
    return f"[{area}] {b['nombre']}" if area else b["nombre"]

class VistaMovimientos(ft.Container):
    def __init__(self, page: ft.Page):
        super().__init__()
        self.page_ref = page
        self.expand = True
        self.datos_cargados = False
        self._cargando = False
        self._picker_masivo = ft.FilePicker(on_result=self._on_masivo_result)
        page.overlay.append(self._picker_masivo)
        self.construir_ui()
        self.content = self._main_content

    def mostrar_snack(self, mensaje, tipo_o_color="success"):
        from utils.ui_helpers import mostrar_snack as _snack
        _snack(self.page_ref, mensaje, tipo_o_color)

    def mostrar_dialogo(self, titulo, mensaje, icono=ft.Icons.INFO, color_icono="blue700", acciones=None):
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(icono, color=color_icono, size=26),
                ft.Text(titulo, weight="bold", size=15),
            ], spacing=10),
            content=ft.Text(mensaje, size=13),
            actions=acciones or [],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        self.page_ref.show_dialog(dlg)
        return dlg

    def cerrar_dialogo(self, dlg):
        self.page_ref.pop_dialog()

    def construir_ui(self):
        solo_lectura = not estado.puede_hacer("movimientos", "ejecutar")

        self.sel_origen = ft.Dropdown(label="Bodega Origen", width=250, options=[])
        self.sel_destino = ft.Dropdown(label="Bodega Destino", width=250, options=[])
        
        self.buscador_sku = BuscadorProducto(
            label="SKU o Nombre", width=180,
            on_seleccionar=self._on_producto_seleccionado,
        )
        self._txt_preview_sku = ft.Text("", size=12, weight="bold")
        self._cont_preview_sku = ft.Container(
            content=self._txt_preview_sku, visible=False,
            padding=ft.Padding(12, 6, 12, 6), border_radius=8,
            bgcolor="blue50", border=ft.Border(left=ft.BorderSide(3, "blue400")),
        )
        self.in_cant = ft.TextField(label="Cantidad", width=100, text_align="center", value="1")
        
        self.btn_traspasar = ft.ElevatedButton(
            "EJECUTAR TRASPASO",
            icon=ft.Icons.SWAP_HORIZ,
            bgcolor="blue800", color="white", height=55,
            on_click=self.procesar_traspaso,
            disabled=solo_lectura
        )

        # ── Traspaso masivo ──────────────────────────────────────────────────
        self._masivo_txt_archivo = ft.Text("Ningún archivo seleccionado", size=12, color="grey500", italic=True, expand=True)
        self._masivo_btn_procesar = ft.ElevatedButton(
            "IMPORTAR Y EJECUTAR", icon=ft.Icons.PLAY_ARROW,
            bgcolor="teal700", color="white", height=45,
            on_click=self.masivo_importar,
            disabled=True,
        )
        self._masivo_archivo_path = [None]

        self._cont_masivo = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.TABLE_CHART, color="teal700", size=22),
                    ft.Text("Traspaso Masivo por Planilla Excel", weight="bold", size=15, color="teal900"),
                ], spacing=8),
                ft.Text(
                    "Descarga la plantilla, completa las filas (SKU, Cantidad, ID Bodega Origen, ID Bodega Destino) "
                    "y luego importa el archivo para ejecutar todos los traspasos en lote.",
                    size=12, color="grey600",
                ),
                ft.Divider(height=8, color="transparent"),
                ft.Row([
                    ft.ElevatedButton(
                        "DESCARGAR PLANTILLA", icon=ft.Icons.DOWNLOAD,
                        bgcolor="grey700", color="white", height=40,
                        on_click=self.masivo_descargar_plantilla,
                    ),
                    ft.ElevatedButton(
                        "SELECCIONAR ARCHIVO .xlsx", icon=ft.Icons.FOLDER_OPEN,
                        bgcolor="blue700", color="white", height=40,
                        on_click=self.masivo_seleccionar_archivo,
                        disabled=solo_lectura,
                    ),
                ], spacing=12, wrap=True),
                ft.Row([
                    ft.Icon(ft.Icons.ATTACH_FILE, color="grey500", size=16),
                    self._masivo_txt_archivo,
                ], spacing=6),
                self._masivo_btn_procesar,
            ], spacing=10),
            padding=ft.Padding(24, 20, 24, 20),
            bgcolor="teal50",
            border_radius=12,
            border=ft.border.all(1, "teal200"),
        )

        self.tabla_historial = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Fecha")),
                ft.DataColumn(ft.Text("Usuario")),
                ft.DataColumn(ft.Text("Detalle del Movimiento")),
            ],
            rows=[]
        )

        from utils.ui_helpers import sandwich

        self.btn_traspasar.tooltip = "Ejecutar traspaso entre bodegas"

        _header = ft.Row([
            ft.Column([
                ft.Text("Movimientos y Traspasos", size=24, weight="bold", color="blue900"),
                ft.Text("Mueve stock entre bodegas físicas o móviles.", color="grey700", size=13),
            ], spacing=2, expand=True),
            ft.IconButton(ft.Icons.REFRESH, tooltip="Actualizar", on_click=self._recargar),
        ])

        _body = ft.Column([
            ft.Container(
                content=ft.Column([
                    ft.Row([
                        self.sel_origen,
                        ft.Icon(ft.Icons.ARROW_FORWARD, color="grey400"),
                        self.sel_destino,
                    ], alignment="center"),
                    ft.Row([
                        self.buscador_sku,
                        self.in_cant,
                        self.btn_traspasar
                    ], alignment="center"),
                    self._cont_preview_sku,
                ], spacing=20),
                padding=30, bgcolor="white", border_radius=12, border=ft.border.all(1, "grey200")
            ),
            ft.Divider(height=10, color="transparent"),
            self._cont_masivo,
            ft.Divider(height=20, color="transparent"),
            ft.Row([
                ft.Text("Últimos Traspasos", weight="bold", size=16, color="grey800"),
                ft.IconButton(ft.Icons.REFRESH, on_click=lambda _: self.cargar_historial(), tooltip="Actualizar historial")
            ]),
            ft.Container(
                content=self.tabla_historial,
                expand=True,
            )
        ], scroll=ft.ScrollMode.AUTO, expand=True)

        self._main_content = sandwich(_header, _body)

    def _on_producto_seleccionado(self, prod):
        stk = prod.get("stock_global") or 0
        color_stk = "red600" if stk == 0 else "orange600" if stk <= 3 else "green700"
        self._txt_preview_sku.value = f"📦  {prod['nombre']}   ·   SKU: {prod['sku']}   ·   Stock: {stk} uds"
        self._txt_preview_sku.color = "blue900"
        self._cont_preview_sku.bgcolor = "blue50"
        self._cont_preview_sku.border = ft.Border(left=ft.BorderSide(3, color_stk))
        self._cont_preview_sku.visible = True
        if self.page_ref: self._cont_preview_sku.update()

    def _poblar_desde_cache(self, app_cache):
        """Aplica datos del cache a los controles y hace content-swap. Sin update()."""
        bodegas = app_cache.mov_bodegas or []
        ops = [ft.dropdown.Option(str(b["id"]), _label_bodega(b)) for b in bodegas]
        self.sel_origen.options = ops
        self.sel_destino.options = ops
        self.tabla_historial.rows = self._filas_desde(app_cache.mov_historial or [])
        self.content = self._main_content
        self.datos_cargados = True
        self._cargando = False

    def inicializar(self):
        if self.datos_cargados or self._cargando:
            return
        self._cargando = True
        hilo(self._cargar_desde_cache)

    def _cargar_desde_cache(self):
        """Espera el evento de cache y renderiza — solo se usa si el cache no estaba listo al navegar."""
        from core.cache import app_cache
        app_cache.mov_ready.wait(timeout=60)
        print(f"[MOV] cache tardó — renderizando desde hilo de espera")
        self._poblar_desde_cache(app_cache)
        if self.page_ref:
            self.page_ref.update()

    @staticmethod
    def _filas_desde(registros):
        import datetime as _dt
        filas = []
        for a in registros:
            fecha_str = "—"
            fecha_raw = a.get("fecha")
            if fecha_raw:
                try:
                    fecha_str = _dt.datetime.fromisoformat(fecha_raw.replace("Z", "+00:00")) \
                        .astimezone().strftime("%d/%m/%Y %H:%M")
                except Exception:
                    fecha_str = str(fecha_raw)[:16]
            detalle = f"SKU: {a['sku']} | {a['cantidad']} uds | {a['ubicacion_origen']} ➡ {a['ubicacion_destino']}"
            filas.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(fecha_str, size=12)),
                ft.DataCell(ft.Text(a.get("usuario", "—"), size=12, weight="bold")),
                ft.DataCell(ft.Text(detalle, size=12, color="grey800")),
            ]))
        return filas

    def _recargar(self, e=None):
        from core.cache import app_cache
        app_cache.mov_ready.clear()
        self.datos_cargados = False
        self._cargando = True
        if self.page_ref:
            self.page_ref.update()
        self._carga_inicial()

    def _carga_inicial(self):
        """Re-obtiene datos desde la BD, actualiza cache y hace content-swap."""
        def _run():
            try:
                from core.cache import app_cache
                sb = get_sb()

                # 1. Bodegas
                bodegas = sb.table("bodegas").select("id, nombre, areas(nombre)").order("id").execute().data or []
                ops = [ft.dropdown.Option(str(b["id"]), _label_bodega(b)) for b in bodegas]
                self.sel_origen.options = ops
                self.sel_destino.options = ops

                # 2. Historial reciente (sin update() aún)
                registros = sb.table("movimientos_inventario").select("*") \
                    .eq("tipo_movimiento", "TRASPASO_INTERNO") \
                    .order("fecha", desc=True).limit(30).execute().data or []
                self.tabla_historial.rows = self._filas_desde(registros)

                # 3. Actualizar cache
                app_cache.mov_bodegas = bodegas
                app_cache.mov_historial = registros
                app_cache.mov_ready.set()

                # 4. Content-swap
                self.content = self._main_content
                self.datos_cargados = True
                self._cargando = False
                if self.page_ref:
                    self.page_ref.update()

            except Exception as e:
                self._cargando = False
                self.mostrar_snack(f"❌ Error al cargar: {e}", "red700")
        hilo(_run)

    def cargar_bodegas(self):
        def _run():
            try:
                res = get_sb().table("bodegas").select("id, nombre, areas(nombre)").order("id").execute()
                bodegas = res.data or []
                ops = [ft.dropdown.Option(str(b["id"]), _label_bodega(b)) for b in bodegas]
                self.sel_origen.options = ops
                self.sel_destino.options = ops
                if self.page_ref:
                    self.sel_origen.update()
                    self.sel_destino.update()
            except Exception as e:
                self.mostrar_snack(f"❌ Error al cargar bodegas: {e}", "red700")
        hilo(_run)

    def procesar_traspaso(self, e):
        origen_id = self.sel_origen.value
        destino_id = self.sel_destino.value
        sku = self.buscador_sku.value.strip()

        if not origen_id or not destino_id:
            self.mostrar_snack("❌ Selecciona ambas bodegas.", "red")
            return
        if not sku:
            self.mostrar_snack("❌ Ingresa un SKU o nombre de producto.", "red")
            return

        if origen_id == destino_id:
            self.mostrar_snack("❌ Origen y destino no pueden ser iguales.", "red")
            return
        try:
            cantidad = int(self.in_cant.value)
            if cantidad <= 0: raise ValueError
        except:
            self.mostrar_snack("❌ Cantidad debe ser un número entero positivo.", "red")
            return

        # Obtener nombres de bodegas para mostrar en el diálogo
        nom_origen = next((o.text for o in self.sel_origen.options if o.key == origen_id), origen_id)
        nom_destino = next((o.text for o in self.sel_destino.options if o.key == destino_id), destino_id)

        def _ejecutar(dlg):
            self.cerrar_dialogo(dlg)
            def _run():
                try:
                    usuario_login = estado.usuario_actual.get("usuario", "")
                    usuario_nombre = estado.usuario_actual.get("nombre", "Usuario Desconocido")
                    ejecutar_traspaso(sku, int(origen_id), int(destino_id), cantidad, usuario_login, usuario_nombre)
                    self.buscador_sku.limpiar()
                    self._cont_preview_sku.visible = False
                    self.in_cant.value = "1"
                    self.page_ref.update()
                    self.cargar_historial()
                    dlg_ok = [None]
                    dlg_ok[0] = self.mostrar_dialogo(
                        "Traspaso ejecutado",
                        f"Se trasladaron {cantidad} unidad(es) del SKU {sku}\nde '{nom_origen}' a '{nom_destino}'.",
                        icono=ft.Icons.CHECK_CIRCLE,
                        color_icono="green600",
                        acciones=[ft.TextButton("Aceptar", on_click=lambda e: self.cerrar_dialogo(dlg_ok[0]))]
                    )
                except ValueError as ve:
                    dlg_err = [None]
                    dlg_err[0] = self.mostrar_dialogo("Error en traspaso", str(ve), icono=ft.Icons.ERROR, color_icono="red700",
                                         acciones=[ft.TextButton("Cerrar", on_click=lambda e: self.cerrar_dialogo(dlg_err[0]))])
                except Exception as ex:
                    dlg_ex = [None]
                    dlg_ex[0] = self.mostrar_dialogo("Error de base de datos", str(ex), icono=ft.Icons.ERROR, color_icono="red700",
                                         acciones=[ft.TextButton("Cerrar", on_click=lambda e: self.cerrar_dialogo(dlg_ex[0]))])
            hilo(_run)

        dlg_confirm = [None]

        def _on_confirmar(e):
            self.cerrar_dialogo(dlg_confirm[0])
            _ejecutar(dlg_confirm[0])

        def _on_cancelar(e):
            self.cerrar_dialogo(dlg_confirm[0])

        dlg_confirm[0] = self.mostrar_dialogo(
            "Confirmar Traspaso",
            f"¿Traspasar  {cantidad} unidad(es) del SKU {sku}\nde '{nom_origen}'  →  '{nom_destino}'?",
            icono=ft.Icons.SWAP_HORIZ,
            color_icono="blue700",
            acciones=[
                ft.TextButton("Cancelar", on_click=_on_cancelar),
                ft.ElevatedButton("Sí, ejecutar", bgcolor="blue800", color="white", on_click=_on_confirmar),
            ]
        )

    # ── Traspaso Masivo ──────────────────────────────────────────────────────
    def masivo_descargar_plantilla(self, e):
        def _run():
            import openpyxl, io, base64 as _b64
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Traspasos"

                headers = ["SKU", "Cantidad", "Bodega_Origen_ID", "Bodega_Destino_ID"]
                fill_h = PatternFill("solid", fgColor="1565C0")
                font_h = Font(bold=True, color="FFFFFF")
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.fill = fill_h; c.font = font_h
                    c.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[get_column_letter(col)].width = 22

                ws.cell(row=2, column=1, value="SKU001")
                ws.cell(row=2, column=2, value=5)
                ws.cell(row=2, column=3, value=1)
                ws.cell(row=2, column=4, value=2)
                for col in range(1, 5):
                    ws.cell(row=2, column=col).font = Font(italic=True, color="888888")

                ws_bod = wb.create_sheet("Bodegas (referencia)")
                ws_bod.cell(row=1, column=1, value="ID").font = Font(bold=True)
                ws_bod.cell(row=1, column=2, value="Nombre").font = Font(bold=True)
                try:
                    res = get_sb().table("bodegas").select("id, nombre").order("id").execute()
                    for i, b in enumerate(res.data or [], start=2):
                        ws_bod.cell(row=i, column=1, value=b["id"])
                        ws_bod.cell(row=i, column=2, value=b["nombre"])
                    ws_bod.column_dimensions["A"].width = 8
                    ws_bod.column_dimensions["B"].width = 30
                except Exception:
                    pass

                # Formato texto en columna SKU para evitar conversión a float
                for row_num in range(2, 502):
                    ws.cell(row=row_num, column=1).number_format = "@"

                _buf = io.BytesIO(); wb.save(_buf)
                import subprocess as _sp, datetime as _dt
                _fname = f"/tmp/PlantillaTraspasos_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                open(_fname, 'wb').write(_buf.getvalue())
                _sp.Popen(['open', _fname])
                self.mostrar_snack("✅ Plantilla generada.", "teal700")
            except Exception as ex:
                self.mostrar_snack(f"❌ Error al crear plantilla: {ex}", "red700")
        hilo(_run)

    def _on_masivo_result(self, e: ft.FilePickerResultEvent):
        import os
        if e.files and e.files[0].path:
            ruta = e.files[0].path
            self._masivo_archivo_path[0] = ruta
            self._masivo_txt_archivo.value = os.path.basename(ruta)
            self._masivo_txt_archivo.color = "blue700"
            self._masivo_btn_procesar.disabled = False
        else:
            self._masivo_archivo_path[0] = None
            self._masivo_txt_archivo.value = "Ningún archivo seleccionado"
            self._masivo_txt_archivo.color = "grey500"
            self._masivo_btn_procesar.disabled = True
        if self.page_ref:
            self._masivo_txt_archivo.update()
            self._masivo_btn_procesar.update()

    def masivo_seleccionar_archivo(self, e):
        self._picker_masivo.pick_files(
            dialog_title="Seleccionar planilla de traspaso masivo",
            allowed_extensions=["xlsx", "xls"],
            allow_multiple=False,
        )

    def masivo_importar(self, e):
        ruta = self._masivo_archivo_path[0]
        if not ruta:
            self.mostrar_snack("❌ Selecciona un archivo primero.", "red700")
            return

        def _run():
            import openpyxl
            try:
                wb = openpyxl.load_workbook(ruta, data_only=True)
                ws = wb.active
            except Exception as ex:
                self.mostrar_snack(f"❌ No se pudo abrir el archivo: {ex}", "red700")
                return

            usuario_login = estado.usuario_actual.get("usuario", "")
            usuario_nombre = estado.usuario_actual.get("nombre", "Usuario")

            # Cargar bodegas válidas para validación
            try:
                res_bod = get_sb().table("bodegas").select("id, nombre").execute()
                bodegas_validas = {str(b["id"]): b["nombre"] for b in (res_bod.data or [])}
            except Exception as ex:
                self.mostrar_snack(f"❌ Error cargando bodegas: {ex}", "red700")
                return

            ok, errores, omitidas = 0, [], []

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                sku = str(row[0]).strip() if row[0] is not None else ""
                cant_raw = row[1]
                origen_raw = row[2]
                destino_raw = row[3]

                # Validaciones básicas
                if not sku:
                    omitidas.append(f"Fila {i}: SKU vacío")
                    continue
                try:
                    cantidad = int(cant_raw)
                    if cantidad <= 0:
                        raise ValueError
                except (TypeError, ValueError):
                    errores.append(f"Fila {i} [{sku}]: Cantidad inválida ({cant_raw!r})")
                    continue
                try:
                    origen_id = int(origen_raw)
                    destino_id = int(destino_raw)
                except (TypeError, ValueError):
                    errores.append(f"Fila {i} [{sku}]: IDs de bodega inválidos ({origen_raw!r}, {destino_raw!r})")
                    continue
                if str(origen_id) not in bodegas_validas:
                    errores.append(f"Fila {i} [{sku}]: Bodega origen {origen_id} no existe")
                    continue
                if str(destino_id) not in bodegas_validas:
                    errores.append(f"Fila {i} [{sku}]: Bodega destino {destino_id} no existe")
                    continue
                if origen_id == destino_id:
                    errores.append(f"Fila {i} [{sku}]: Origen y destino iguales")
                    continue

                try:
                    ejecutar_traspaso(sku, origen_id, destino_id, cantidad, usuario_login, usuario_nombre)
                    ok += 1
                except ValueError as ve:
                    errores.append(f"Fila {i} [{sku}]: {ve}")
                except Exception as ex:
                    errores.append(f"Fila {i} [{sku}]: Error inesperado — {ex}")

            # Registrar auditoría de la operación masiva
            if ok > 0:
                registrar_auditoria(usuario_nombre, "TRASPASO MASIVO",
                                    f"{ok} traspasos ejecutados desde planilla xlsx")
                self.cargar_historial()

            # Resetear selección
            self._masivo_archivo_path[0] = None
            self._masivo_txt_archivo.value = "Ningún archivo seleccionado"
            self._masivo_txt_archivo.color = "grey500"
            self._masivo_btn_procesar.disabled = True
            if self.page_ref:
                self._masivo_txt_archivo.update()
                self._masivo_btn_procesar.update()

            # Mostrar resultado en diálogo
            lineas_error = "\n".join(errores[:20])
            if len(errores) > 20:
                lineas_error += f"\n… y {len(errores) - 20} errores más"
            lineas_omit = "\n".join(omitidas[:10])

            resumen = f"✅ Traspasos exitosos: {ok}"
            if errores:
                resumen += f"\n❌ Errores: {len(errores)}"
            if omitidas:
                resumen += f"\n⏭  Filas omitidas: {len(omitidas)}"
            if errores:
                resumen += f"\n\nDetalle errores:\n{lineas_error}"
            if omitidas:
                resumen += f"\n\nFilas omitidas:\n{lineas_omit}"

            icono = ft.Icons.CHECK_CIRCLE if not errores else ft.Icons.WARNING_AMBER
            color_icono = "green600" if not errores else "orange700"
            dlg_ref = [None]
            dlg_ref[0] = self.mostrar_dialogo(
                "Resultado Traspaso Masivo",
                resumen,
                icono=icono,
                color_icono=color_icono,
                acciones=[ft.ElevatedButton("Aceptar", bgcolor="blue800", color="white",
                                             on_click=lambda _: self.cerrar_dialogo(dlg_ref[0]))]
            )
        hilo(_run)

    # ── Historial ────────────────────────────────────────────────────────────
    def cargar_historial(self):
        def _run():
            try:
                res = get_sb().table("movimientos_inventario").select("*").eq("tipo_movimiento", "TRASPASO_INTERNO").order("fecha", desc=True).limit(30).execute()
                registros = res.data or []

                import datetime
                filas = []
                for a in registros:
                    fecha_str = "—"
                    fecha_raw = a.get("fecha")
                    if fecha_raw:
                        try:
                            fecha_str = datetime.datetime.fromisoformat(fecha_raw.replace("Z", "+00:00")).astimezone().strftime("%d/%m/%Y %H:%M")
                        except:
                            fecha_str = str(fecha_raw)[:16]

                    detalle = f"SKU: {a['sku']} | {a['cantidad']} uds | {a['ubicacion_origen']} ➡ {a['ubicacion_destino']}"

                    filas.append(ft.DataRow(cells=[
                        ft.DataCell(ft.Text(fecha_str, size=12)),
                        ft.DataCell(ft.Text(a.get("usuario", "—"), size=12, weight="bold")),
                        ft.DataCell(ft.Text(detalle, size=12, color="grey800")),
                    ]))

                self.tabla_historial.rows = filas
                if self.page_ref:
                    self.page_ref.update()
            except Exception as ex:
                self.mostrar_snack(f"❌ Error al cargar historial: {ex}", "red700")
        hilo(_run)
